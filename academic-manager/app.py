from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, make_response, send_file
from flask_mysqldb import MySQL
import MySQLdb.cursors
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import hashlib
import secrets
import os
import io
from io import StringIO, BytesIO, TextIOWrapper
import csv
import json
import openpyxl

app = Flask(__name__)
   
app.secret_key = '40509aa19fec5150daf28c9cb55cc7157f2b5b4d11103de3c1be0950ad331d8d'
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'academic_manager'

mysql = MySQL(app)

UPLOAD_FOLDER = 'static/profile_pics'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

UPLOAD_FOLDER2 = 'static/assignment_uploads'
ALLOWED_EXTENSIONS2 = {'pdf', 'docx', 'txt', 'zip'}
app.config['UPLOAD_FOLDER2'] = UPLOAD_FOLDER2

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
           
#helper functions
@app.template_filter('timeformat')
def timeformat_filter(hour):
    if hour > 12:
        return f"{hour - 12}:00 PM"
    elif hour == 12:
        return "12:00 PM"
    else:
        return f"{hour}:00 AM"
    
@app.template_filter('twelve_hour')
def twelve_hour(hour):
    """Convert 24-hour to 12-hour format"""
    return hour if hour <= 12 else hour - 12

@app.template_filter('am_pm')
def am_pm(hour):
    """Return AM/PM for a given hour"""
    return 'AM' if hour < 12 else 'PM'

@app.context_processor
def inject_datetime():
    return dict(datetime=datetime)

def get_current_semester():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT * FROM semesters WHERE is_current = TRUE LIMIT 1")
    return cursor.fetchone()

def get_student_id(user_id):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT student_id FROM students WHERE user_id = %s", (user_id,))
    student = cursor.fetchone()
    return student['student_id'] if student else None

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

@app.route("/api/current-periods")
def get_current_periods():
    if 'loggedin' in session:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("""
            SELECT enroll_start, enroll_end, drop_start, drop_end 
            FROM semesters 
            WHERE is_current = TRUE LIMIT 1
        """)
        period = cursor.fetchone()
        today = datetime.now().date()
        return jsonify({
            'enroll_active': period and period['enroll_start'] and period['enroll_end'] and 
                            today >= period['enroll_start'] and today <= period['enroll_end'],
            'drop_active': period and period['drop_start'] and period['drop_end'] and 
                          today >= period['drop_start'] and today <= period['drop_end'],
            'enroll_period': {
                'start': period['enroll_start'].strftime('%b %d, %Y') if period and period['enroll_start'] else None,
                'end': period['enroll_end'].strftime('%b %d, %Y') if period and period['enroll_end'] else None},
            'drop_period': {
                'start': period['drop_start'].strftime('%b %d, %Y') if period and period['drop_start'] else None,
                'end': period['drop_end'].strftime('%b %d, %Y') if period and period['drop_end'] else None}})
    return jsonify({'error': 'Unauthorized'}), 401

#authentication routes
@app.route('/')
@app.route('/login', methods=['GET', 'POST'])
def login():
    message = ''
    if request.method == 'POST' and 'email' in request.form and 'password' in request.form:
        email = request.form['email']
        password = request.form['password']
        hashed_password = hash_password(password)
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM user WHERE email = %s AND password = %s', (email, hashed_password, ))
        user = cursor.fetchone()
        if user:
            session['loggedin'] = True
            session['userid'] = user['id']
            session['name'] = user['first_name']
            session['last_name']=user['last_name']
            session['email'] = user['email']
            session['role'] = user['role']
            message = 'Logged in successfully!'
            if user['role'] == 'student':
                return redirect(url_for('student_dashboard'))
            elif user['role'] == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('admin_dashboard'))
        else:
            message = 'Please enter correct email / password!'
    return render_template('login.html', message=message)

@app.route('/logout')
def logout():
    session.pop('loggedin', None)
    session.pop('userid', None)
    session.pop('email', None)
    session.pop('role', None)
    return redirect(url_for('login'))

#password reset routes
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST' and 'email' in request.form:
        email = request.form['email']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT id FROM user WHERE email = %s', (email,))
        user = cursor.fetchone()
        if user:
            token = secrets.token_urlsafe(32)
            expires_at = datetime.now() + timedelta(hours=1)
            cursor.execute('INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (%s, %s, %s)', (user['id'], token, expires_at))
            mysql.connection.commit()
            #example link reset
            reset_link = url_for('reset_password', token=token, _external=True)
            message = f'A password reset link has been sent to your email. For demo: {reset_link}'
            return render_template('forgot_password.html', message=message)
        message = 'If this email exists, a reset link has been sent.'
        return render_template('forgot_password.html', message=message)
    return render_template('forgot_password.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute(
        'SELECT user_id FROM password_reset_tokens WHERE token = %s AND used = FALSE AND expires_at > NOW()',
        (token,)
    )
    valid_token = cursor.fetchone()
    
    if not valid_token:
        return render_template('reset_password.html', valid=False)
    
    if request.method == 'POST' and 'password' in request.form and 'confirm_password' in request.form:
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        if password != confirm_password:
            return render_template('reset_password.html', valid=True, message='Passwords do not match')
        hashed_password = hash_password(password)
        cursor.execute('UPDATE user SET password = %s WHERE id = %s',(hashed_password, valid_token['user_id']))
        cursor.execute('UPDATE password_reset_tokens SET used = TRUE WHERE token = %s',(token,))
        mysql.connection.commit()
        return render_template('reset_password.html', valid=True, success=True)
    return render_template('reset_password.html', valid=True)

@app.route("/student/dashboard")
def student_dashboard():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        current_semester = get_current_semester()
        
        # Distinct courses grouped by course and section
        cursor.execute("""
            SELECT 
                c.course_id,
                c.title, 
                c.credits, 
                s.section_id, 
                s.section_number,
                CONCAT(u.first_name, ' ', u.last_name) AS instructor_name
            FROM enrollments e
            JOIN sections s ON e.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN user u ON s.instructor_id = u.id
            WHERE e.student_id = %s 
            AND s.semester_id = %s 
            AND e.status = 'registered'
            GROUP BY c.course_id, s.section_number, CONCAT(u.first_name, ' ', u.last_name), s.semester_id
            ORDER BY c.title, s.section_number
        """, (student_id, current_semester['semester_id'] if current_semester else None))
        current_courses = cursor.fetchall()
        
        # Upcoming assignments
        cursor.execute("""
            SELECT a.title, a.due_date, c.title as course_title, 
                   s.section_id, s.section_number
            FROM assignments a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            JOIN enrollments e ON s.section_id = e.section_id
            WHERE e.student_id = %s AND e.status = 'registered' 
            AND a.due_date > NOW()
            ORDER BY a.due_date ASC
            LIMIT 5
        """, (student_id,))
        upcoming_assignments = cursor.fetchall()
        
        # Recent announcements from all enrolled courses
        cursor.execute("""
            SELECT a.announcement_id, a.title, a.message, a.created_at,
                   c.title as course_title, s.section_number,
                   CONCAT(u.first_name, ' ', u.last_name) as instructor_name
            FROM announcements a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            JOIN user u ON a.teacher_id = u.id
            JOIN enrollments e ON s.section_id = e.section_id
            WHERE e.student_id = %s
            ORDER BY a.created_at DESC
            LIMIT 5
        """, (student_id,))
        recent_announcements = cursor.fetchall()
        
        return render_template("student/dashboard.html",
                             current_semester=current_semester,
                             current_courses=current_courses,
                             upcoming_assignments=upcoming_assignments,
                             recent_announcements=recent_announcements)
    return redirect(url_for('login'))

@app.route("/student/announcements")
def student_announcements():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get all announcements grouped by course/section
        cursor.execute("""
            SELECT 
                a.announcement_id, 
                a.title, 
                a.message, 
                a.created_at,
                c.course_id,
                c.title as course_title, 
                s.section_id,
                s.section_number,
                CONCAT(u.first_name, ' ', u.last_name) as instructor_name
            FROM announcements a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            JOIN user u ON a.teacher_id = u.id
            JOIN enrollments e ON s.section_id = e.section_id
            WHERE e.student_id = %s AND e.status="registered"
            ORDER BY c.title, s.section_number, a.created_at DESC
        """, (student_id,))
        announcements = cursor.fetchall()
        
        # Group announcements by course/section
        announcements_by_course = {}
        for announcement in announcements:
            key = f"{announcement['course_title']} - Section {announcement['section_number']}"
            if key not in announcements_by_course:
                announcements_by_course[key] = {
                    'course_id': announcement['course_id'],
                    'section_id': announcement['section_id'],
                    'instructor': announcement['instructor_name'],
                    'announcements': []
                }
            announcements_by_course[key]['announcements'].append(announcement)
        
        return render_template("student/announcements.html",
                             announcements_by_course=announcements_by_course)
    return redirect(url_for('login'))

#student profile
@app.route("/student/profile")
def student_profile():
    if 'loggedin' in session and session['role'] == 'student':
        user_id = session['userid']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM user WHERE id = %s', (user_id,))
        user = cursor.fetchone()
        cursor.execute('SELECT * FROM students WHERE user_id = %s', (user_id,))
        student = cursor.fetchone()
        if student:
            user.update(student)
        return render_template("student/profile.html", user=user)
    return redirect(url_for('login'))

@app.route("/upload_profile_pic", methods=['POST'])
def upload_profile_pic():
    if 'loggedin' not in session or session['role'] != 'student':
        return redirect(url_for('login'))
    
    if 'profile_pic' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('student_profile'))
    
    file = request.files['profile_pic']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('student_profile'))
    
    if file and allowed_file(file.filename):
        user_id = session['userid']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get current profile pic filename to delete it later
        cursor.execute('SELECT profile_pic FROM students WHERE user_id = %s', (user_id,))
        student = cursor.fetchone()
        old_filename = student['profile_pic'] if student and student['profile_pic'] else None
        filename = secure_filename(f"student_{user_id}_{file.filename}")
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        cursor.execute('UPDATE students SET profile_pic = %s WHERE user_id = %s', (filename, user_id))
        mysql.connection.commit()
        #deleting old file
        if old_filename and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], old_filename)):
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], old_filename))
        
        flash('Profile picture updated successfully!', 'success')
    else:
        flash('Allowed file types are: png, jpg, jpeg, gif', 'danger')
    
    return redirect(url_for('student_profile'))

@app.route("/student/attendance")
def student_attendance():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        current_semester = get_current_semester()
        
        if not current_semester:
            return render_template("student/attendance.html",
                                attendance_data={},
                                current_semester=None)
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get all course sections with attendance data
        cursor.execute("""
            SELECT 
                c.course_id,
                c.title,
                s.section_id,
                s.section_number,
                CONCAT(u.first_name, ' ', u.last_name) as instructor_name,
                s.semester_id,
                COUNT(a.attendance_id) as total_classes,
                SUM(CASE WHEN a.status IN ('present', 'excused') THEN 1 ELSE 0 END) as present_classes
            FROM enrollments e
            JOIN sections s ON e.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN user u ON s.instructor_id = u.id
            LEFT JOIN attendance a ON a.section_id = s.section_id 
                AND a.course_id = c.course_id 
                AND a.student_id = %s
            WHERE e.student_id = %s 
            AND s.semester_id = %s
            AND e.status = 'registered'
            GROUP BY c.course_id, s.section_number, s.instructor_id, s.semester_id
            ORDER BY c.title, s.section_number
        """, (student_id, student_id, current_semester['semester_id']))
        attendance_data = {}
        for course_section in cursor.fetchall():
            #calculate attendance percentage
            attendance_percentage = round((course_section['present_classes'] / course_section['total_classes'] * 100), 2) if course_section['total_classes'] > 0 else 0
            # Get detailed attendance records
            cursor.execute("""
                SELECT date, status, duration
                FROM attendance
                WHERE student_id = %s 
                AND section_id = %s
                AND course_id = %s
                ORDER BY date DESC
            """, (student_id, course_section['section_id'], course_section['course_id']))
            records = cursor.fetchall()
            #create a unique key for this course-section
            key = f"{course_section['course_id']}-{course_section['section_number']}"
            attendance_data[key] = {
                'title': course_section['title'],
                'section_number': course_section['section_number'],
                'instructor_name': course_section['instructor_name'] or 'TBA',
                'records': records,
                'stats': {
                    'total_classes': course_section['total_classes'],
                    'present_classes': course_section['present_classes'],
                    'attendance_percentage': attendance_percentage}}
        return render_template("student/attendance.html",
                            attendance_data=attendance_data,
                            current_semester=current_semester)
    return redirect(url_for('login'))

@app.route("/student/courses")
def student_courses():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        current_semester = get_current_semester()
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        now = datetime.now().date()
        #get enrollment period
        cursor.execute("SELECT enroll_start, enroll_end FROM semesters WHERE is_current = TRUE")
        enroll_period = cursor.fetchone()
        current_enroll_period = (enroll_period['enroll_start'], enroll_period['enroll_end']) if enroll_period else None
        #get drop period
        cursor.execute("SELECT drop_start, drop_end FROM semesters WHERE is_current = TRUE")
        drop_period = cursor.fetchone()
        current_drop_period = (drop_period['drop_start'], drop_period['drop_end']) if drop_period else None
        #get available courses
        cursor.execute("""
        SELECT 
            c.course_id as code,
            c.title, 
            c.credits,
            MIN(s.section_id) as section_id,
            s.section_number,
            u.first_name, 
            u.last_name,
            s.instructor_id,  # ADD THIS
            COUNT(DISTINCT s.day) as meeting_days,
            (SELECT COUNT(*) 
            FROM enrollments e 
            WHERE e.section_id = MIN(s.section_id)
            AND e.status = 'registered') as enrolled,
            s.max_capacity
        FROM sections s
        JOIN courses c ON s.course_id = c.course_id
        LEFT JOIN user u ON s.instructor_id = u.id
        WHERE s.semester_id = %s
        GROUP BY c.course_id, c.title, c.credits, 
                s.section_number, u.first_name, u.last_name, s.instructor_id, s.max_capacity
        ORDER BY c.title, s.section_number
    """, (current_semester['semester_id'],))
        
        available_courses = cursor.fetchall()
        #get enrolled courses
        cursor.execute("""
        SELECT 
            e.enrollment_id, 
            c.course_id as code,
            c.title, 
            c.credits, 
            s.section_id, 
            s.section_number,
            u.first_name, 
            u.last_name,
            s.instructor_id  # ADD THIS
        FROM enrollments e
        JOIN sections s ON e.section_id = s.section_id
        JOIN courses c ON s.course_id = c.course_id
        LEFT JOIN user u ON s.instructor_id = u.id
        WHERE e.student_id = %s AND s.semester_id = %s AND e.status = 'registered'
        GROUP BY s.section_number,c.course_id,u.first_name,u.last_name,s.semester_id,s.instructor_id
        ORDER BY c.title
    """, (student_id, current_semester['semester_id']))
        
        enrolled_courses = cursor.fetchall()
        #get all schedules for available courses
        available_schedules = {}
        cursor.execute("""
            SELECT 
                s.section_id,
                s.section_number,
                s.course_id,
                s.instructor_id,
                s.semester_id,
                s.day,
                TIME_FORMAT(s.time_start, '%%h:%%i %%p') as time_start,
                TIME_FORMAT(s.time_end, '%%h:%%i %%p') as time_end
            FROM sections s
            WHERE s.semester_id = %s
            ORDER BY s.section_id, 
                FIELD(s.day, 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'),
                s.time_start
        """, (current_semester['semester_id'],))
        
        for schedule in cursor.fetchall():
            key = (schedule['section_number'], schedule['course_id'], 
                   schedule['instructor_id'], schedule['semester_id'])
            if key not in available_schedules:
                available_schedules[key] = []
            schedule_str = f"{schedule['day'].capitalize()} {schedule['time_start']}-{schedule['time_end']}"
            available_schedules[key].append(schedule_str)
        
        #get all schedules for enrolled courses
        enrolled_schedules = {}
        cursor.execute("""
            SELECT 
                s.section_id,
                s.section_number,
                s.course_id,
                s.instructor_id,
                s.semester_id,
                s.day,
                TIME_FORMAT(s.time_start, '%%h:%%i %%p') as time_start,
                TIME_FORMAT(s.time_end, '%%h:%%i %%p') as time_end
            FROM enrollments e
            JOIN sections s ON e.section_id = s.section_id
            WHERE e.student_id = %s AND s.semester_id = %s AND e.status = 'registered'
            ORDER BY s.section_id, 
                FIELD(s.day, 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'),
                s.time_start
        """, (student_id, current_semester['semester_id']))
        
        for schedule in cursor.fetchall():
            key = (schedule['section_number'], schedule['course_id'], 
                   schedule['instructor_id'], schedule['semester_id'])
            if key not in enrolled_schedules:
                enrolled_schedules[key] = []
            schedule_str = f"{schedule['day'].capitalize()} {schedule['time_start']}-{schedule['time_end']}"
            enrolled_schedules[key].append(schedule_str)
        
        cursor.close()
        return render_template(
            "student/courses.html",
            current_enroll_period=current_enroll_period,
            current_drop_period=current_drop_period,
            now=now,
            current_semester=current_semester,
            available_courses=available_courses,
            enrolled_courses=enrolled_courses,
            available_schedules=available_schedules,
            enrolled_schedules=enrolled_schedules)
    return redirect(url_for('login'))

@app.route("/student/courses/enroll", methods=['POST'])
def enroll_course():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        section_id = request.form.get('section_id')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            #get the section details including grouping information
            cursor.execute("""
                SELECT 
                    s.section_id,
                    s.section_number,
                    s.course_id,
                    s.instructor_id,
                    s.semester_id,
                    s.max_capacity,
                    s.day,
                    s.time_start,
                    s.time_end,
                    COUNT(CASE WHEN e.status = 'registered' THEN 1 END) as current_enrollment
                FROM sections s
                LEFT JOIN enrollments e ON s.section_id = e.section_id
                WHERE s.section_id = %s
                GROUP BY s.section_id
            """, (section_id,))
            target_section = cursor.fetchone()
            if not target_section:
                return jsonify({'success': False, 'message': 'Section not found'})
            #check if already enrolled in any section with same grouping
            cursor.execute("""
                SELECT e.* 
                FROM enrollments e
                JOIN sections s ON e.section_id = s.section_id
                WHERE e.student_id = %s 
                AND e.status = 'registered'
                AND s.section_number = %s
                AND s.course_id = %s
                AND s.instructor_id = %s
                AND s.semester_id = %s
            """, (student_id,
                target_section['section_number'],
                target_section['course_id'],
                target_section['instructor_id'],
                target_section['semester_id']))
            existing_enrollment = cursor.fetchone()
            if existing_enrollment:
                return jsonify({
                    'success': False, 
                    'message': f'Already enrolled in another section of this course with the same instructor'})
            #check capacity
            if target_section['current_enrollment'] >= target_section['max_capacity']:
                return jsonify({'success': False, 'message': 'Course is full'})
            #get all schedules for this section group
            cursor.execute("""
                SELECT 
                    s.day,
                    s.time_start,
                    s.time_end
                FROM sections s
                WHERE s.section_number = %s
                AND s.course_id = %s
                AND s.instructor_id = %s
                AND s.semester_id = %s
                ORDER BY 
                    FIELD(s.day, 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'),
                    s.time_start
            """, (target_section['section_number'],
                target_section['course_id'],
                target_section['instructor_id'],
                target_section['semester_id']))
            all_schedules = cursor.fetchall()
            #get all existing enrollments with their schedules
            cursor.execute("""
                SELECT 
                    s.day,
                    s.time_start,
                    s.time_end,
                    c.title
                FROM enrollments e
                JOIN sections s ON e.section_id = s.section_id
                JOIN courses c ON s.course_id = c.course_id
                WHERE e.student_id = %s 
                AND e.status = 'registered'
            """, (student_id,))
            existing_courses = cursor.fetchall()
            #convert timedelta to time for comparison
            def timedelta_to_time(td):
                seconds = td.seconds
                hours = seconds // 3600
                minutes = (seconds % 3600) // 60
                return datetime.strptime(f"{hours:02d}:{minutes:02d}", "%H:%M").time()
            #check for time conflicts with all schedules in the section group
            for schedule in all_schedules:
                new_day = schedule['day'].lower()
                new_start = timedelta_to_time(schedule['time_start'])
                new_end = timedelta_to_time(schedule['time_end'])
                for course in existing_courses:
                    existing_day = course['day'].lower()
                    existing_start = timedelta_to_time(course['time_start'])
                    existing_end = timedelta_to_time(course['time_end'])
                    if (existing_day == new_day and 
                        ((new_start < existing_end and new_end > existing_start))):
                        return jsonify({
                            'success': False, 
                            'message': f"Time conflict with {course['title']} ({course['day']} {existing_start.strftime('%I:%M %p')}-{existing_end.strftime('%I:%M %p')})"})
            #get all distinct courses the student is enrolled in for this semester
            cursor.execute("""
                SELECT DISTINCT e.course_id
                FROM enrollments e
                JOIN sections s ON e.section_id = s.section_id
                WHERE e.student_id = %s 
                AND e.status = 'registered'
                AND s.semester_id = %s
            """, (student_id, target_section['semester_id']))
            enrolled_courses = cursor.fetchall()
            #calculate total credits by summing credits for each distinct course
            total_credits = 0
            if enrolled_courses:
                course_ids = [str(course['course_id']) for course in enrolled_courses]
                cursor.execute(f"""
                    SELECT SUM(credits) as total
                    FROM courses
                    WHERE course_id IN ({','.join(course_ids)})
                """)
                total_credits = cursor.fetchone()['total'] or 0
            print(f"total user prev enrolled: {total_credits}")
            #get the semester's max credits
            cursor.execute("""
                SELECT max_credits 
                FROM semesters 
                WHERE semester_id = %s
            """, (target_section['semester_id'],))
            max_credits = cursor.fetchone()['max_credits']
            print(f"max credits: {max_credits}")
            #get the new course's credits
            cursor.execute("""
                SELECT credits FROM courses 
                WHERE course_id = %s
            """, (target_section['course_id'],))
            new_course_credits = cursor.fetchone()['credits']
            print(f"new course credits: {new_course_credits}")
            print(total_credits + new_course_credits)
            #check if adding this course would exceed max credits
            if (total_credits + new_course_credits) > max_credits:
                return jsonify({
                    'success': False, 
                    'message': f'Enrollment would exceed max credits ({total_credits + new_course_credits}/{max_credits})'})
            mysql.connection.begin()
            # Enroll student in all sections with the same grouping
            cursor.execute("""
                INSERT INTO enrollments 
                (student_id, section_id, course_id, status) 
                SELECT %s, s.section_id, s.course_id, 'registered'
                FROM sections s
                WHERE s.section_number = %s
                AND s.course_id = %s
                AND s.instructor_id = %s
                AND s.semester_id = %s
                ON DUPLICATE KEY UPDATE status = 'registered'
            """, (
                student_id,
                target_section['section_number'],
                target_section['course_id'],
                target_section['instructor_id'],
                target_section['semester_id']))
            mysql.connection.commit()
            return jsonify({
                'success': True, 
                'message': f'Successfully enrolled in all sessions for this course section'})
        except Exception as e:
            mysql.connection.rollback()
            return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    return jsonify({'success': False, 'message': 'Unauthorized'}), 401

@app.route("/student/courses/drop", methods=['POST'])
def drop_course():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        enrollment_id = request.form.get('enrollment_id')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            #get the section details for this enrollment
            cursor.execute("""
                SELECT 
                    s.section_number,
                    s.course_id,
                    s.instructor_id,
                    s.semester_id
                FROM enrollments e
                JOIN sections s ON e.section_id = s.section_id
                WHERE e.enrollment_id = %s
                AND e.student_id = %s
                AND e.status = 'registered'
            """, (enrollment_id, student_id))
            enrollment_info = cursor.fetchone()
            if not enrollment_info:
                return jsonify({'success': False, 'message': 'Active enrollment not found'})

            mysql.connection.begin()
            #drop all enrollments in the same course group
            cursor.execute("""
                UPDATE enrollments e
                JOIN sections s ON e.section_id = s.section_id
                SET e.status = 'dropped'
                WHERE e.student_id = %s
                AND e.status = 'registered'
                AND s.section_number = %s
                AND s.course_id = %s
                AND s.instructor_id = %s
                AND s.semester_id = %s
            """, (student_id,
                enrollment_info['section_number'],
                enrollment_info['course_id'],
                enrollment_info['instructor_id'],
                enrollment_info['semester_id']))
            affected_rows = cursor.rowcount
            mysql.connection.commit()
            if affected_rows > 0:
                return jsonify({
                    'success': True, 
                    'message': f'Successfully dropped all {affected_rows} sessions for this course'})
            else:
                return jsonify({
                    'success': False, 
                    'message': 'No active enrollments found for this course group'})
        except Exception as e:
            mysql.connection.rollback()
            return jsonify({
                'success': False, 
                'message': f'Error dropping course: {str(e)}'}), 500
    return jsonify({'success': False, 'message': 'Unauthorized'}), 401

@app.route("/student/courses/<int:section_id>")
def course_details(section_id):
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        current_semester = get_current_semester()
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify enrollment in this specific section
        cursor.execute('SELECT * FROM enrollments WHERE student_id = %s AND section_id = %s AND status = "registered"', 
                      (student_id, section_id))
        if not cursor.fetchone():
            return redirect(url_for('student_courses'))
        
        # Get course info and grouping information for this specific section
        cursor.execute("""
            SELECT 
                c.*,
                s.section_id,
                s.section_number,
                s.instructor_id,
                s.semester_id,
                CONCAT(s.day, ' ', TIME_FORMAT(s.time_start, '%%h:%%i %%p'), ' - ', TIME_FORMAT(s.time_end, '%%h:%%i %%p')) AS schedule,
                u.first_name, 
                u.last_name, 
                u.email as instructor_email
            FROM sections s
            JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN user u ON s.instructor_id = u.id
            WHERE s.section_id = %s
        """, (section_id,))
        course = cursor.fetchone()
        if not course:
            return redirect(url_for('student_courses'))
        
        # Get all assignments for this course group
        cursor.execute("""
            SELECT 
                a.*, 
                s.grades AS grade,
                s.submit_date AS submission_date,
                sec.section_id
            FROM assignments a
            JOIN sections sec ON a.section_id = sec.section_id
            LEFT JOIN assignment_submission s 
                ON a.assignment_id = s.assignment_id 
                AND s.student_id = %s
            WHERE sec.section_number = %s
                AND sec.course_id = %s
                AND sec.instructor_id = %s
                AND sec.semester_id = %s
            ORDER BY a.due_date
        """, (student_id,
            course['section_number'],
            course['course_id'],
            course['instructor_id'],
            course['semester_id']))
        assignments = cursor.fetchall()
        
        # Get exam results with new schema
        cursor.execute("""
            SELECT 
                exam_id,
                course_id,
                section_id,
                semester_id,
                mid1,
                mid2,
                project,
                quizzes,
                assignments,
                final_exam,
                final_result,
                created_at,
                updated_at
            FROM exam_results
            WHERE student_id = %s
                AND course_id = %s
                AND section_id = %s
                AND semester_id = %s
            ORDER BY created_at DESC
            LIMIT 1
        """, (student_id,
            course['course_id'],
            section_id,
            course['semester_id']))
        
        exam_results = cursor.fetchone()
        
        # Get all attendance for this course group
        cursor.execute("""
            SELECT date, status 
            FROM attendance 
            WHERE student_id = %s
                AND section_id IN (
                    SELECT section_id FROM sections 
                    WHERE section_number = %s
                    AND course_id = %s
                    AND instructor_id = %s
                    AND semester_id = %s
                )
            ORDER BY date DESC
            LIMIT 10
        """, (student_id,
            course['section_number'],
            course['course_id'],
            course['instructor_id'],
            course['semester_id']))
        attendance = cursor.fetchall()
        
        return render_template("student/course_details.html",
                             course=course,
                             assignments=assignments,
                             exam_results=exam_results,
                             attendance=attendance,
                             current_date=datetime.now())
    return redirect(url_for('login'))

@app.route("/student/exams")
def student_exams():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        current_semester = get_current_semester()
        
        if not current_semester:
            return render_template("student/exams.html",
                                courses=[],
                                exam_results={},
                                current_semester=None)

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get all course groups the student is enrolled in for current semester
        cursor.execute("""
            SELECT 
                c.course_id,
                c.title,
                s.section_number,
                s.instructor_id,
                s.semester_id,
                MIN(s.section_id) as primary_section_id
            FROM enrollments e
            JOIN sections s ON e.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            WHERE e.student_id = %s 
            AND s.semester_id = %s
            AND e.status = 'registered'
            GROUP BY c.course_id, c.title, s.section_number, s.instructor_id, s.semester_id
            ORDER BY c.title, s.section_number
        """, (student_id, current_semester['semester_id']))
        
        course_groups = cursor.fetchall()
        exam_results = {}

        # Get exam results for each course
        for group in course_groups:
            cursor.execute("""
                SELECT 
                    exam_id,
                    student_id,
                    course_id,
                    section_id,
                    semester_id,
                    mid1,
                    mid2,
                    project,
                    quizzes,
                    assignments,
                    final_exam,
                    final_result,
                    created_at,
                    updated_at
                FROM exam_results
                WHERE student_id = %s
                AND course_id = %s
                AND section_id = %s
                AND semester_id = %s
            """, (student_id, 
                 group['course_id'],
                 group['primary_section_id'],
                 group['semester_id']))
            
            exams = cursor.fetchall()
            
            # Calculate course average if exams exist
            course_average = None
            if exams:
                # Calculate based on available components
                components = []
                total_weight = 0
                total_score = 0
                
                # Check each component and add to calculation if exists
                for exam in exams:
                    if exam['mid1'] is not None:
                        components.append(('Midterm 1', exam['mid1'], 20))  # Example weight 20%
                    if exam['mid2'] is not None:
                        components.append(('Midterm 2', exam['mid2'], 20))  # Example weight 20%
                    if exam['project'] is not None:
                        components.append(('Project', exam['project'], 15))  # Example weight 15%
                    if exam['quizzes'] is not None:
                        components.append(('Quizzes', exam['quizzes'], 15))  # Example weight 15%
                    if exam['assignments'] is not None:
                        components.append(('Assignments', exam['assignments'], 10))  # Example weight 10%
                    if exam['final_exam'] is not None:
                        components.append(('Final Exam', exam['final_exam'], 20))  # Example weight 20%
                
                # Calculate weighted average
                if components:
                    total_score = sum(score * weight for (name, score, weight) in components)
                    total_weight = sum(weight for (name, score, weight) in components)
                    
                    if total_weight > 0:
                        course_average = round((total_score / total_weight) * 100, 2)
            
            exam_results[group['course_id']] = {
                'exams': exams,
                'average': course_average,
                'section_number': group['section_number'],
                'instructor_id': group['instructor_id'],
                'final_result': exams[0]['final_result'] if exams else None
            }

        return render_template("student/exams.html",
                            courses=course_groups,
                            exam_results=exam_results,
                            current_semester=current_semester)
    
    return redirect(url_for('login'))

@app.route("/student/assignments")
def student_assignments():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get all courses the student is enrolled in
        cursor.execute("""
            SELECT DISTINCT c.course_id, c.title
            FROM enrollments e
            JOIN sections s ON e.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            WHERE e.student_id = %s AND e.status = 'registered'
            ORDER BY c.course_id
        """, (student_id,))
        courses = cursor.fetchall()
        #get pending assignments (not submitted yet)
        cursor.execute("""
            SELECT a.assignment_id, a.title, a.due_date, a.max_points,
                c.course_id, s.section_id
            FROM assignments a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            JOIN enrollments e ON s.section_id = e.section_id
            WHERE e.student_id = %s 
            AND a.assignment_id NOT IN (
                SELECT assignment_id FROM assignment_submission 
                WHERE student_id = %s
            )
            AND a.due_date > NOW()
            ORDER BY a.due_date ASC
        """, (student_id, student_id))
        pending_assignments = cursor.fetchall()
        #get submitted but not graded assignments
        cursor.execute("""
            SELECT a.assignment_id, a.title, a.due_date, a.max_points,
                c.course_id, s.section_id,
                sub.submit_date, sub.file_path, sub.grades as grade
            FROM assignments a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            JOIN assignment_submission sub ON a.assignment_id = sub.assignment_id
            WHERE sub.student_id = %s
            AND sub.grades IS NULL
            ORDER BY sub.submit_date DESC
        """, (student_id,))
        submitted_assignments = cursor.fetchall()
        #get graded assignments
        cursor.execute("""
            SELECT a.assignment_id, a.title, a.due_date, a.max_points,
                c.course_id, s.section_id,
                sub.submit_date, sub.grades as grade, sub.feedback
            FROM assignments a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            JOIN assignment_submission sub ON a.assignment_id = sub.assignment_id
            WHERE sub.student_id = %s
            AND sub.grades IS NOT NULL
            ORDER BY sub.submit_date DESC
        """, (student_id,))
        graded_assignments = cursor.fetchall()
        return render_template("student/assignments.html",
                             courses=courses,
                             pending_assignments=pending_assignments,
                             submitted_assignments=submitted_assignments,
                             graded_assignments=graded_assignments,
                             current_date=datetime.now())
    return redirect(url_for('login'))

@app.route("/student/assignments/submit", methods=['POST'])
def submit_assignment():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        assignment_id = request.form.get('assignment_id')
        comments = request.form.get('comments', '')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            #check if already submitted
            cursor.execute('SELECT file_path FROM assignment_submission WHERE assignment_id = %s AND student_id = %s',(assignment_id, student_id))
            existing = cursor.fetchone()
            #handle file upload
            file_path = None
            if 'file' in request.files:
                file = request.files['file']
                if file.filename:
                    #delete old file if it exists
                    if existing and existing['file_path']:
                        try:
                            #extract relative path (remove '/static/' prefix)
                            rel_path = existing['file_path'].replace('/static/', '')
                            old_file = os.path.join(app.root_path, 'static', rel_path)
                            if os.path.exists(old_file):
                                os.remove(old_file)
                        except Exception as e:
                            print(f"Warning: Could not delete old file - {e}")
                    #save new file with basic filename sanitization
                    clean_name = ''.join(c for c in file.filename if c.isalnum() or c in (' ', '.', '_'))
                    filename = f"assignment_{assignment_id}_{student_id}_{clean_name}"
                    upload_dir = os.path.join(app.root_path, 'static', 'uploads')
                    os.makedirs(upload_dir, exist_ok=True)
                    file_path = os.path.join('uploads', filename)
                    file.save(os.path.join(upload_dir, filename))
                    file_path = f"/static/{file_path}"  #format for database storage
            if existing:
                cursor.execute('UPDATE assignment_submission SET submit_date=NOW(), comments=%s, file_path=%s WHERE assignment_id=%s AND student_id=%s',(comments, file_path, assignment_id, student_id) )
            else:
                cursor.execute('INSERT INTO assignment_submission (assignment_id, student_id, comments, file_path) VALUES (%s, %s, %s, %s)',(assignment_id, student_id, comments, file_path))
            mysql.connection.commit()
            return jsonify({'success': True, 'message': 'Assignment submitted'})
        except Exception as e:
            mysql.connection.rollback()
            return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    return jsonify({'success': False, 'message': 'Unauthorized'}), 401

@app.route("/student/transcript")
def student_transcript():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get current semester
        current_semester = get_current_semester()
        if not current_semester:
            return render_template("student/transcript.html",
                                current_grades=[],
                                semesters=[],
                                semester_grades=[],
                                selected_semester=None,
                                semester_gpa=None,
                                transcript_records=[],
                                overall_gpa=None)
        #get current semester grades
        cursor.execute("""
            SELECT c.course_id as code, c.title, c.credits, 
                   t.grade as letter_grade, 
                   t.points as numeric_grade,
                   t.type
            FROM transcript t
            JOIN courses c ON t.course_id = c.course_id
            WHERE t.student_id = %s AND t.semester_id = %s
            ORDER BY c.title
        """, (student_id, current_semester['semester_id']))
        current_grades = cursor.fetchall()
        #get all semesters with grades
        cursor.execute("""
            SELECT DISTINCT s.semester_id, s.name as semester_name
            FROM transcript t
            JOIN semesters s ON t.semester_id = s.semester_id
            WHERE t.student_id = %s
            ORDER BY s.start_date DESC
        """, (student_id,))
        semesters = cursor.fetchall()
        #get selected semester grades
        selected_semester = request.args.get('semester', current_semester['semester_id'])
        cursor.execute("""
            SELECT c.course_id as code, c.title, c.credits, 
                   t.grade as letter_grade, 
                   t.points as numeric_grade,
                   t.type
            FROM transcript t
            JOIN courses c ON t.course_id = c.course_id
            WHERE t.student_id = %s AND t.semester_id = %s
            ORDER BY c.title
        """, (student_id, selected_semester))
        semester_grades = cursor.fetchall()
        #calculate GPA for selected semester with NULL handling
        cursor.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN t.points IS NOT NULL THEN c.credits ELSE 0 END), 0) as total_credits,
                COALESCE(SUM(t.points * c.credits), 0) as weighted_points
            FROM transcript t
            JOIN courses c ON t.course_id = c.course_id
            WHERE t.student_id = %s AND t.semester_id = %s
        """, (student_id, selected_semester))
        gpa_data = cursor.fetchone()
        
        semester_gpa = None
        if gpa_data and gpa_data['total_credits'] > 0:
            semester_gpa = round(gpa_data['weighted_points'] / gpa_data['total_credits'], 2)
        #get complete transcript records
        cursor.execute("""
            SELECT s.name as semester_name, 
                   c.course_id as code,
                   c.title, 
                   c.credits, 
                   t.grade, 
                   t.points, 
                   t.type
            FROM transcript t
            JOIN courses c ON t.course_id = c.course_id
            JOIN semesters s ON t.semester_id = s.semester_id
            WHERE t.student_id = %s
            ORDER BY s.start_date DESC, c.title
        """, (student_id,))
        transcript_records = cursor.fetchall()
        #calculate overall GPA with NULL handling
        cursor.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN t.points IS NOT NULL THEN c.credits ELSE 0 END), 0) as total_credits,
                COALESCE(SUM(t.points * c.credits), 0) as weighted_points
            FROM transcript t
            JOIN courses c ON t.course_id = c.course_id
            WHERE t.student_id = %s
        """, (student_id,))
        overall_gpa_data = cursor.fetchone()
        
        overall_gpa = None
        if overall_gpa_data and overall_gpa_data['total_credits'] > 0:
            overall_gpa = round(overall_gpa_data['weighted_points'] / overall_gpa_data['total_credits'], 2)
        
        #map the field names to what the template expects
        for grade in current_grades:
            grade['grade'] = grade['letter_grade']
            grade['points'] = grade['numeric_grade'] 
        for grade in semester_grades:
            grade['grade'] = grade['letter_grade']
            grade['points'] = grade['numeric_grade']
        
        return render_template("student/transcript.html",
                            current_grades=current_grades,
                            semesters=semesters,
                            semester_grades=semester_grades,
                            selected_semester=int(selected_semester) if selected_semester else None,
                            semester_gpa=semester_gpa,
                            transcript_records=transcript_records,
                            overall_gpa=overall_gpa)
    return redirect(url_for('login'))

#financial Records
@app.route("/student/finances")
def student_finances():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get all semesters with financial records
        cursor.execute("""
            SELECT DISTINCT s.semester_id, s.name
            FROM financial_record f
            JOIN semesters s ON f.semester_id = s.semester_id
            WHERE f.student_id = %s
            ORDER BY s.start_date DESC
        """, (student_id,))
        semesters = cursor.fetchall()
        #get selected semester records(default to current semester)
        default_semester = get_current_semester()
        if isinstance(default_semester, dict):
            default_semester = default_semester['semester_id']
        selected_semester = request.args.get('semester', default_semester)
        cursor.execute("""
            SELECT 
                f.semester_id,
                s.name AS description,
                f.amount,
                f.due_date AS transaction_date,
                f.payment_method,
                f.status
            FROM financial_record f
            JOIN semesters s ON f.semester_id = s.semester_id
            WHERE f.student_id = %s AND f.semester_id = %s
            ORDER BY f.due_date DESC
        """, (student_id, selected_semester))
        records = cursor.fetchall()
        #calculate totals
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN status = 'paid' THEN amount ELSE 0 END) as paid_amount,
                SUM(CASE WHEN status = 'pending' THEN amount ELSE 0 END) as pending_amount
            FROM financial_record
            WHERE student_id = %s AND semester_id = %s
        """, (student_id, selected_semester))
        totals = cursor.fetchone()
        return render_template("student/finances.html",
                            semesters=semesters,
                            records=records,
                            selected_semester=int(selected_semester),
                            totals=totals)
    return redirect(url_for('login'))

#internships and career opportunities
@app.route("/student/internships")
def student_internships():
    if 'loggedin' in session and session['role'] == 'student':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get available internships
        cursor.execute("""
            SELECT * FROM internship 
            WHERE is_active = TRUE AND application_deadline >= CURDATE()
            ORDER BY posted_date DESC
        """)
        internships = cursor.fetchall()
        return render_template("student/internships.html",
                            internships=internships)
    return redirect(url_for('login'))

#library
@app.route("/student/library")
def student_library():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #search functionality
        search_query = request.args.get('q', '')
        if search_query:
            cursor.execute("""
                SELECT * FROM books 
                WHERE title LIKE %s OR author LIKE %s OR isbn = %s
                ORDER BY title
            """, (f'%{search_query}%', f'%{search_query}%', search_query))
        else:
            cursor.execute("SELECT * FROM books ORDER BY title LIMIT 50")
        books = cursor.fetchall()
        #get current transactions
        cursor.execute("""
            SELECT t.*, b.title, b.author
            FROM book_reserve t
            JOIN books b ON t.book_id = b.book_id
            WHERE t.student_id = %s AND t.status = 'active'
            ORDER BY t.due_date
        """, (student_id,))
        current_loans = cursor.fetchall()
        return render_template("student/library.html",
                            books=books,
                            current_loans=current_loans,
                            search_query=search_query)
    return redirect(url_for('login'))

@app.route("/student/library/borrow", methods=['POST'])
def borrow_book():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        book_id = request.form.get('book_id')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #check book availability
        cursor.execute('SELECT available_copies FROM books WHERE book_id = %s',(book_id,))
        book = cursor.fetchone()
        if not book or book['available_copies'] <= 0:
            return jsonify({'success': False, 'message': 'Book not available for borrowing'})
        #check if student already has this book
        cursor.execute('SELECT * FROM book_reserve WHERE book_id = %s AND student_id = %s AND status = "active"',(book_id, student_id))
        if cursor.fetchone():
            return jsonify({'success': False, 'message': 'You already have this book checked out'})
        #calculate due date(2 weeks from now)
        due_date = datetime.now() + timedelta(days=14)
        #create transaction
        cursor.execute('INSERT INTO book_reserve (book_id, student_id, type, due_date) VALUES (%s, %s, "borrow", %s)',(book_id, student_id, due_date))
        #update available copies
        cursor.execute('UPDATE books SET available_copies = available_copies - 1 WHERE book_id = %s',(book_id,))
        mysql.connection.commit()
        return jsonify({'success': True, 'message': 'Book borrowed successfully'})
    return jsonify({'success': False, 'message': 'Unauthorized'}), 401

@app.route("/student/schedule")
def student_schedule():
    if 'loggedin' in session and session['role'] == 'student':
        student_id = get_student_id(session['userid'])
        current_semester = get_current_semester()
        if not current_semester:
            return render_template("student/schedule.html",
                                   schedule={},
                                   current_semester=None)
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get all enrolled courses with their schedule info
        cursor.execute("""
            SELECT 
                c.title,
                s.day,
                s.time_start,
                s.time_end,
                s.section_number,
                CONCAT(u.first_name, ' ', u.last_name) as instructor
            FROM enrollments e
            JOIN sections s ON e.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN user u ON s.instructor_id = u.id
            WHERE e.student_id = %s 
            AND s.semester_id = %s
            AND e.status = 'registered'
            ORDER BY s.day, s.time_start
        """, (student_id, current_semester['semester_id']))
        courses = cursor.fetchall()
        #create schedule grid (8AM-5PM)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        hours = list(range(8, 18))#8 to 5
        schedule = {day: {hour: None for hour in hours} for day in days}
        for course in courses:
            day = course['day'].capitalize()
            try:
                #convert time format from timedelta
                start_time = (datetime.min + course['time_start']).time()
                end_time = (datetime.min + course['time_end']).time()
                #format the time display
                formatted_start = start_time.strftime('%I:%M %p').lstrip('0')
                formatted_end = end_time.strftime('%I:%M %p').lstrip('0')
                course_time = f"{formatted_start} - {formatted_end}"
                start_hour = start_time.hour
                end_hour = end_time.hour
                duration_hours = end_hour - start_hour
                #mark all affected hours
                for hour_slot in range(start_hour, end_hour):
                    if hour_slot in schedule[day]:
                        schedule[day][hour_slot] = {
                            'title': course['title'],
                            'section': course['section_number'],
                            'instructor': course['instructor'],
                            'time': course_time,
                            'is_first': (hour_slot == start_hour),
                            'duration': duration_hours,
                            'is_continuation': (hour_slot > start_hour)}
            except Exception as e:
                print(f"Error processing schedule for course {course['title']}: {e}")
                continue
        print(schedule)
        return render_template("student/schedule.html",
                               schedule=schedule,
                               days=days,
                               hours=hours,
                               current_semester=current_semester)
    return redirect(url_for('login'))

#admin
@app.route("/admin/dashboard")
def admin_dashboard():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        first_name = session.get('first_name', 'Admin')
        last_name = session.get('last_name', 'User')
        user_name = f"{first_name} {last_name}"
        cursor.execute("SELECT * FROM semesters WHERE is_current = TRUE LIMIT 1")
        current_semester = cursor.fetchone()
        stats = {
            'students_count': 0,
            'courses_count': 0,
            'overdue_books': 0,
            'teachers_count': 0,
            'active_courses': 0}
        cursor.execute("SELECT COUNT(*) as count FROM students")
        stats['students_count'] = cursor.fetchone()['count']
        cursor.execute("SELECT COUNT(*) as count FROM courses")
        stats['courses_count'] = cursor.fetchone()['count']
        cursor.execute("SELECT COUNT(*) as count FROM user WHERE role = 'teacher'")
        stats['teachers_count'] = cursor.fetchone()['count']
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM book_reserve 
            WHERE status = 'active' AND due_date < CURDATE()
        """)
        stats['overdue_books'] = cursor.fetchone()['count']
        if current_semester:
            cursor.execute("""
                SELECT COUNT(DISTINCT course_id) as count 
                FROM sections 
                WHERE semester_id = %s
            """, (current_semester['semester_id'],))
            stats['active_courses'] = cursor.fetchone()['count']
        #get current semester courses
        current_courses = []
        if current_semester:
            cursor.execute("""
                SELECT c.title, 
                       COUNT(DISTINCT s.section_id) as section_count,
                       COUNT(e.enrollment_id) as enrollment_count,
                       CONCAT(u.first_name, ' ', u.last_name) as instructor_name
                FROM sections s
                JOIN courses c ON s.course_id = c.course_id
                LEFT JOIN enrollments e ON s.section_id = e.section_id AND e.status = 'registered'
                LEFT JOIN user u ON s.instructor_id = u.id
                WHERE s.semester_id = %s
                GROUP BY c.course_id, instructor_name
                ORDER BY c.title
                LIMIT 5
            """, (current_semester['semester_id'],))
            current_courses = cursor.fetchall()
        return render_template("admin/dashboard.html",
                            current_semester=current_semester,
                            stats=stats,
                            current_courses=current_courses,
                            user=user_name)
    return redirect(url_for('login'))

#user management routes
@app.route("/admin/users")
def admin_users():
    if 'loggedin' in session and session['role'] == 'admin':
        role_filter = request.args.get('role', 'all')
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        per_page = 10
        query = """
            SELECT * FROM user 
            WHERE 1=1
        """
        count_query = """
            SELECT COUNT(*) as total FROM user 
            WHERE 1=1
        """
        params = []
        #role filter
        if role_filter != 'all':
            query += " AND role = %s"
            count_query += " AND role = %s"
            params.append(role_filter) 
        #search functionality
        if search:
            search = search.strip()  # Clean up the search term
            search_term = f"%{search}%"
            #search by ID (exact match), email, first name, last name, or full name
            query += """
                AND (id = %s OR 
                    email LIKE %s OR 
                    first_name LIKE %s OR 
                    last_name LIKE %s OR 
                    CONCAT(first_name, ' ', last_name) LIKE %s)"""
            count_query += """
                AND (id = %s OR 
                    email LIKE %s OR 
                    first_name LIKE %s OR 
                    last_name LIKE %s OR 
                    CONCAT(first_name, ' ', last_name) LIKE %s)"""
            try:
                id_search = int(search)
            except ValueError:
                id_search = 0 
            params.extend([id_search, search_term, search_term, search_term, search_term])
        # Ordering and pagination
        query += " ORDER BY last_name, first_name LIMIT %s OFFSET %s"
        #get total count
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        #get paginated results
        params.extend([per_page, (page-1)*per_page])
        cursor.execute(query, params)
        users = cursor.fetchall()
        return render_template("admin/users/list.html",
                            users=users,
                            role_filter=role_filter,
                            search=search,
                            pagination={
                                'page': page,
                                'per_page': per_page,
                                'total': total,
                                'pages': (total + per_page - 1) // per_page})
    return redirect(url_for('login'))

@app.route("/admin/users/create", methods=['GET', 'POST'])
def create_user():
    if 'loggedin' in session and session['role'] == 'admin':
        if request.method == 'POST':
            email = request.form['email']
            password = hash_password(request.form['password'])
            first_name = request.form['first_name']
            last_name = request.form['last_name']
            role = request.form['role']
            is_active = 'is_active' in request.form
            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            try:
                cursor.execute("INSERT INTO user (email, password, first_name, last_name, role, is_active) VALUES (%s, %s, %s, %s, %s, %s)",(email, password, first_name, last_name, role, is_active))
                mysql.connection.commit()
                if role == 'student':
                    user_id = cursor.lastrowid
                    cursor.execute("INSERT INTO students (user_id, program, enrollment_date) VALUES (%s, %s, %s)",(user_id, 'bachelors', datetime.now().date()))
                    mysql.connection.commit()
                flash('User created successfully!', 'success')
                return redirect(url_for('admin_users'))
            except MySQLdb.IntegrityError as e:
                mysql.connection.rollback()
                flash('Email already exists!', 'danger')
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating user: {str(e)}', 'danger')
        return render_template("admin/users/form.html", user=None)
    return redirect(url_for('login'))

@app.route("/admin/users/<int:user_id>/edit", methods=['GET', 'POST'])
def edit_user(user_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            email = request.form['email']
            first_name = request.form['first_name']
            last_name = request.form['last_name']
            role = request.form['role']
            is_active = 'is_active' in request.form
            try:
                cursor.execute("UPDATE user SET email=%s, first_name=%s, last_name=%s, role=%s, is_active=%s WHERE id=%s",(email, first_name, last_name, role, is_active, user_id))
                mysql.connection.commit()
                flash('User updated successfully!', 'success')
                return redirect(url_for('admin_users'))
            except MySQLdb.IntegrityError as e:
                mysql.connection.rollback()
                flash('Email already exists!', 'danger')
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating user: {str(e)}', 'danger')
        cursor.execute("SELECT * FROM user WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            flash('User not found!', 'danger')
            return redirect(url_for('admin_users'))
        return render_template("admin/users/form.html", user=user)
    return redirect(url_for('login'))

@app.route("/admin/users/<int:user_id>/delete", methods=['POST'])
def delete_user(user_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            cursor.execute("DELETE FROM user WHERE id = %s", (user_id,))
            mysql.connection.commit()
            flash('User deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting user: {str(e)}', 'danger')
        return redirect(url_for('admin_users'))
    return redirect(url_for('login'))

@app.route("/admin/users/<int:user_id>/reset-password", methods=['POST'])
def reset_user_password(user_id):
    if 'loggedin' in session and session['role'] == 'admin':
        new_password = hash_password(request.form['new_password'])
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            cursor.execute( "UPDATE user SET password = %s WHERE id = %s",(new_password, user_id))
            mysql.connection.commit()
            flash('Password reset successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error resetting password: {str(e)}', 'danger')
        return redirect(url_for('edit_user', user_id=user_id))
    return redirect(url_for('login'))

#course management routes
@app.route("/admin/courses")
def admin_courses():
    if 'loggedin' in session and session['role'] == 'admin':
        search = request.args.get('search', '')
        department_id = request.args.get('department_id', 'all')
        page = request.args.get('page', 1, type=int)
        per_page = 10
        course_query = """
            SELECT c.*, d.name as department_name 
            FROM courses c
            LEFT JOIN departments d ON c.department_id = d.department_id
            WHERE 1=1"""
        count_query = "SELECT COUNT(*) as total FROM courses WHERE 1=1"
        params = []
        if search:
            course_query += " AND (c.title LIKE %s OR c.course_id LIKE %s)"
            count_query += " AND (title LIKE %s OR course_id LIKE %s)"
            params.extend([f"%{search}%", f"%{search}%"])
        if department_id != 'all':
            course_query += " AND c.department_id = %s"
            count_query += " AND department_id = %s"
            params.append(department_id)
        #get total count
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        #get paginated results
        course_query += " ORDER BY c.title LIMIT %s OFFSET %s"
        params.extend([per_page, (page-1)*per_page])
        cursor.execute(course_query, params)
        courses = cursor.fetchall()
        #get departments for filter dropdown
        cursor.execute("SELECT * FROM departments ORDER BY name")
        departments = cursor.fetchall()
        return render_template("admin/courses/list.html",
                            courses=courses,
                            departments=departments,
                            search=search,
                            department_id=department_id,
                            pagination={
                                'page': page,
                                'per_page': per_page,
                                'total': total,
                                'pages': (total + per_page - 1) // per_page})
    return redirect(url_for('login'))

@app.route("/admin/courses/create", methods=['GET', 'POST'])
def create_course():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            title = request.form['title']
            credits = request.form['credits']
            department_id = request.form['department_id'] or None
            syllabus_url = request.form.get('syllabus_url', '')
            content_url = request.form.get('content_url', '')
            
            try:
                # Insert basic course info
                cursor.execute(
                    "INSERT INTO courses (title, credits, department_id, syllabus, content_file_path) VALUES (%s, %s, %s, %s, %s)",
                    (title, credits, department_id, syllabus_url, content_url)
                )
                course_id = cursor.lastrowid
                
                # Handle syllabus file upload if present
                if 'syllabus_file' in request.files and request.files['syllabus_file'].filename:
                    file = request.files['syllabus_file']
                    filename = f"course_{course_id}_syllabus_{secure_filename(file.filename)}"
                    upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'syllabi')
                    os.makedirs(upload_dir, exist_ok=True)
                    file_path = os.path.join(upload_dir, filename)
                    file.save(file_path)
                    cursor.execute(
                        "UPDATE courses SET syllabus = %s WHERE course_id = %s",
                        (f"/static/uploads/syllabi/{filename}", course_id)
                    )
                
                # Handle content file upload if present
                if 'content_file' in request.files and request.files['content_file'].filename:
                    file = request.files['content_file']
                    filename = f"course_{course_id}_content_{secure_filename(file.filename)}"
                    upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'content')
                    os.makedirs(upload_dir, exist_ok=True)
                    file_path = os.path.join(upload_dir, filename)
                    file.save(file_path)
                    cursor.execute(
                        "UPDATE courses SET content_file_path = %s WHERE course_id = %s",
                        (f"/static/uploads/content/{filename}", course_id)
                    )
                
                mysql.connection.commit()
                flash('Course created successfully!', 'success')
                return redirect(url_for('admin_courses'))
                
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating course: {str(e)}', 'danger')
        
        # Get departments for dropdown
        cursor.execute("SELECT * FROM departments ORDER BY name")
        departments = cursor.fetchall()
        return render_template("admin/courses/form.html", 
                            course=None, 
                            departments=departments)
    return redirect(url_for('login'))

@app.route("/admin/courses/<int:course_id>/edit", methods=['GET', 'POST'])
def edit_course(course_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            title = request.form['title']
            credits = request.form['credits']
            department_id = request.form['department_id'] or None
            syllabus_url = request.form.get('syllabus_url', '')
            content_url = request.form.get('content_url', '')
            
            try:
                # First update with URL values (may be overwritten by file uploads)
                cursor.execute(
                    "UPDATE courses SET title=%s, credits=%s, department_id=%s, syllabus=%s, content_file_path=%s WHERE course_id=%s",
                    (title, credits, department_id, syllabus_url, content_url, course_id)
                )
                
                # Handle syllabus file upload if present
                if 'syllabus_file' in request.files and request.files['syllabus_file'].filename:
                    file = request.files['syllabus_file']
                    # Delete old syllabus file if it was an uploaded file
                    cursor.execute("SELECT syllabus FROM courses WHERE course_id = %s", (course_id,))
                    old_syllabus = cursor.fetchone()['syllabus']
                    if old_syllabus and old_syllabus.startswith('/static/uploads/syllabi/'):
                        try:
                            os.remove(os.path.join(app.root_path, old_syllabus[1:]))
                        except Exception as e:
                            print(f"Error deleting old syllabus file: {e}")
                    # Save new file
                    filename = f"course_{course_id}_syllabus_{secure_filename(file.filename)}"
                    upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'syllabi')
                    os.makedirs(upload_dir, exist_ok=True)
                    file_path = os.path.join(upload_dir, filename)
                    file.save(file_path)
                    cursor.execute(
                        "UPDATE courses SET syllabus = %s WHERE course_id = %s",
                        (f"/static/uploads/syllabi/{filename}", course_id)
                    )
                
                # Handle content file upload if present
                if 'content_file' in request.files and request.files['content_file'].filename:
                    file = request.files['content_file']
                    # Delete old content file if it was an uploaded file
                    cursor.execute("SELECT content_file_path FROM courses WHERE course_id = %s", (course_id,))
                    old_content = cursor.fetchone()['content_file_path']
                    if old_content and old_content.startswith('/static/uploads/content/'):
                        try:
                            os.remove(os.path.join(app.root_path, old_content[1:]))
                        except Exception as e:
                            print(f"Error deleting old content file: {e}")
                    # Save new file
                    filename = f"course_{course_id}_content_{secure_filename(file.filename)}"
                    upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'content')
                    os.makedirs(upload_dir, exist_ok=True)
                    file_path = os.path.join(upload_dir, filename)
                    file.save(file_path)
                    cursor.execute(
                        "UPDATE courses SET content_file_path = %s WHERE course_id = %s",
                        (f"/static/uploads/content/{filename}", course_id)
                    )
                
                mysql.connection.commit()
                flash('Course updated successfully!', 'success')
                return redirect(url_for('admin_courses'))
                
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating course: {str(e)}', 'danger')
        
        # Get course data
        cursor.execute("SELECT * FROM courses WHERE course_id = %s", (course_id,))
        course = cursor.fetchone()
        if not course:
            flash('Course not found!', 'danger')
            return redirect(url_for('admin_courses')) 
        
        # Get departments for dropdown
        cursor.execute("SELECT * FROM departments ORDER BY name")
        departments = cursor.fetchall()
        return render_template("admin/courses/form.html", 
                            course=course, 
                            departments=departments)
    return redirect(url_for('login'))

@app.route("/admin/courses/<int:course_id>/delete", methods=['POST'])
def delete_course(course_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            #get file path to delete
            cursor.execute("SELECT content_file_path FROM courses WHERE course_id = %s", (course_id,))
            file_path = cursor.fetchone()['content_file_path']
            #deleting course
            cursor.execute("DELETE FROM courses WHERE course_id = %s", (course_id,))
            mysql.connection.commit()
            #delete associated file if exists
            if file_path:
                try:
                    os.remove(os.path.join(app.root_path, 'static', file_path[1:]))
                except Exception as e:
                    print(f"Error deleting file: {e}")
            flash('Course deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting course: {str(e)}', 'danger')
        return redirect(url_for('admin_courses'))
    return redirect(url_for('login'))

#department management routes
@app.route("/admin/departments")
def admin_departments():
    if 'loggedin' in session and session['role'] == 'admin':
        search = request.args.get('search', '')
        query = "SELECT * FROM departments WHERE 1=1"
        params = []
        if search:
            query += " AND (name LIKE %s OR code LIKE %s)"
            params.extend([f"%{search}%", f"%{search}%"])
        query += " ORDER BY name"
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(query, params)
        departments = cursor.fetchall()
        return render_template("admin/departments/list.html",
                            departments=departments,
                            search=search)
    return redirect(url_for('login'))

@app.route("/admin/departments/create", methods=['GET', 'POST'])
def create_department():
    if 'loggedin' in session and session['role'] == 'admin':
        if request.method == 'POST':
            name = request.form['name']
            code = request.form['code']
            description = request.form.get('description', '')
            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            try:
                cursor.execute("INSERT INTO departments (name, code, description) VALUES (%s, %s, %s)",(name, code, description))
                mysql.connection.commit()
                flash('Department created successfully!', 'success')
                return redirect(url_for('admin_departments'))
            except MySQLdb.IntegrityError as e:
                mysql.connection.rollback()
                flash('Department code already exists!', 'danger')
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating department: {str(e)}', 'danger')
        return render_template("admin/departments/form.html", department=None)
    return redirect(url_for('login'))

@app.route("/admin/departments/<int:department_id>/edit", methods=['GET', 'POST'])
def edit_department(department_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            name = request.form['name']
            code = request.form['code']
            description = request.form.get('description', '')
            try:
                cursor.execute("UPDATE departments SET name=%s, code=%s, description=%s WHERE department_id=%s",(name, code, description, department_id))
                mysql.connection.commit()
                flash('Department updated successfully!', 'success')
                return redirect(url_for('admin_departments'))
            except MySQLdb.IntegrityError as e:
                mysql.connection.rollback()
                flash('Department code already exists!', 'danger')
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating department: {str(e)}', 'danger')
        cursor.execute("SELECT * FROM departments WHERE department_id = %s", (department_id,))
        department = cursor.fetchone()
        if not department:
            flash('Department not found!', 'danger')
            return redirect(url_for('admin_departments'))
        return render_template("admin/departments/form.html", department=department)
    return redirect(url_for('login'))

@app.route("/admin/departments/<int:department_id>/delete", methods=['POST'])
def delete_department(department_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            #check if department has courses
            cursor.execute("SELECT COUNT(*) as course_count FROM courses WHERE department_id = %s", (department_id,))
            if cursor.fetchone()['course_count'] > 0:
                flash('Cannot delete department with associated courses!', 'danger')
                return redirect(url_for('admin_departments'))
            cursor.execute("DELETE FROM departments WHERE department_id = %s", (department_id,))
            mysql.connection.commit()
            flash('Department deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting department: {str(e)}', 'danger')
        return redirect(url_for('admin_departments'))
    return redirect(url_for('login'))

#semester management routes
@app.route("/admin/semesters")
def admin_semesters():
    if 'loggedin' in session and session['role'] == 'admin':
        search = request.args.get('search', '')
        query = "SELECT * FROM semesters WHERE 1=1"
        params = []
        if search:
            query += " AND name LIKE %s"
            params.append(f"%{search}%")
        query += " ORDER BY start_date DESC"
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(query, params)
        semesters = cursor.fetchall()
        return render_template("admin/semesters/list.html",
                            semesters=semesters,
                            search=search)
    return redirect(url_for('login'))

@app.route("/admin/semesters/create", methods=['GET', 'POST'])
def create_semester():
    if 'loggedin' in session and session['role'] == 'admin':
        if request.method == 'POST':
            name = request.form['name']
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            max_credits = request.form['max_credits']
            is_current = 'is_current' in request.form
            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            try:
                #if setting as current,first unset any existing current semester
                if is_current:
                    cursor.execute("UPDATE semesters SET is_current = FALSE")
                cursor.execute("INSERT INTO semesters (name, start_date, end_date, max_credits, is_current) VALUES (%s, %s, %s, %s, %s)",(name, start_date, end_date, max_credits, is_current))
                mysql.connection.commit()
                flash('Semester created successfully!', 'success')
                return redirect(url_for('admin_semesters'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating semester: {str(e)}', 'danger')
        return render_template("admin/semesters/form.html", semester=None)
    return redirect(url_for('login'))

@app.route("/admin/semesters/<int:semester_id>/edit", methods=['GET', 'POST'])
def edit_semester(semester_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            name = request.form['name']
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            max_credits = request.form['max_credits']
            is_current = 'is_current' in request.form
            try:
                #if setting as current,first unset any existing current semester
                if is_current:
                    cursor.execute("UPDATE semesters SET is_current = FALSE")
                cursor.execute("UPDATE semesters SET name=%s, start_date=%s, end_date=%s, max_credits=%s, is_current=%s WHERE semester_id=%s",(name, start_date, end_date, max_credits, is_current, semester_id))
                mysql.connection.commit()
                flash('Semester updated successfully!', 'success')
                return redirect(url_for('admin_semesters'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating semester: {str(e)}', 'danger')
        cursor.execute("SELECT * FROM semesters WHERE semester_id = %s", (semester_id,))
        semester = cursor.fetchone()
        if not semester:
            flash('Semester not found!', 'danger')
            return redirect(url_for('admin_semesters')) 
        return render_template("admin/semesters/form.html", semester=semester)
    return redirect(url_for('login'))

@app.route("/admin/semesters/<int:semester_id>/delete", methods=['POST'])
def delete_semester(semester_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            #check if semester has sections
            cursor.execute("SELECT COUNT(*) as section_count FROM sections WHERE semester_id = %s", (semester_id,))
            if cursor.fetchone()['section_count'] > 0:
                flash('Cannot delete semester with associated sections!', 'danger')
                return redirect(url_for('admin_semesters'))
            cursor.execute("DELETE FROM semesters WHERE semester_id = %s", (semester_id,))
            mysql.connection.commit()
            flash('Semester deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting semester: {str(e)}', 'danger')
        return redirect(url_for('admin_semesters'))
    return redirect(url_for('login'))

@app.route("/admin/semesters/<int:semester_id>/set-current", methods=['POST'])
def set_current_semester(semester_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            cursor.execute("UPDATE semesters SET is_current = FALSE")
            cursor.execute("UPDATE semesters SET is_current = TRUE WHERE semester_id = %s",(semester_id,))
            mysql.connection.commit()
            flash('Current semester updated successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error updating current semester: {str(e)}', 'danger')
        return redirect(url_for('admin_semesters'))
    return redirect(url_for('login'))

#section management routes
@app.route("/admin/sections")
def admin_sections():
    if 'loggedin' in session and session['role'] == 'admin':
        semester_id = request.args.get('semester_id', '')
        course_id = request.args.get('course_id', '')
        instructor_id = request.args.get('instructor_id', '')
        page = request.args.get('page', 1, type=int)
        per_page = 10
        section_query = """
            SELECT s.section_id, s.section_number, s.course_id, s.instructor_id, s.semester_id,
                   c.title as course_title, 
                   CONCAT(u.first_name, ' ', u.last_name) as instructor_name,
                   s.max_capacity,
                   COUNT(e.enrollment_id) as enrolled,
                   (SELECT COUNT(*) FROM sections s2 
                    WHERE s2.section_number = s.section_number 
                    AND s2.course_id = s.course_id 
                    AND s2.instructor_id = s.instructor_id 
                    AND s2.semester_id = s.semester_id) as total_meetings
            FROM sections s
            JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN user u ON s.instructor_id = u.id
            LEFT JOIN enrollments e ON s.section_id = e.section_id AND e.status = 'registered'
            WHERE 1=1"""
        count_query = "SELECT COUNT(DISTINCT CONCAT(section_number, course_id, instructor_id, semester_id)) as total FROM sections WHERE 1=1"
        params = []
        if semester_id:
            section_query += " AND s.semester_id = %s"
            count_query += " AND semester_id = %s"
            params.append(semester_id)  
        if course_id:
            section_query += " AND s.course_id = %s"
            count_query += " AND course_id = %s"
            params.append(course_id)
        if instructor_id:
            section_query += " AND s.instructor_id = %s"
            count_query += " AND instructor_id = %s"
            params.append(instructor_id)
        #group by section grouping
        section_query += " GROUP BY s.section_number, s.course_id, s.instructor_id, s.semester_id"
        #get total count
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        #get paginated results
        section_query += " ORDER BY c.title, s.section_number LIMIT %s OFFSET %s"
        params.extend([per_page, (page-1)*per_page])
        cursor.execute(section_query, params)
        sections = cursor.fetchall()
        #get all meetings for each section
        for section in sections:
            cursor.execute("""
                SELECT day, TIME_FORMAT(time_start, '%%H:%%i') as time_start, 
                       TIME_FORMAT(time_end, '%%H:%%i') as time_end
                FROM sections 
                WHERE section_number = %s AND course_id = %s 
                  AND instructor_id <=> %s AND semester_id = %s
                ORDER BY FIELD(day, 'monday', 'tuesday', 'wednesday', 
                              'thursday', 'friday', 'saturday', 'sunday')
            """, (section['section_number'], section['course_id'], 
                  section['instructor_id'], section['semester_id']))
            section['meetings'] = cursor.fetchall()
        #get search options
        cursor.execute("SELECT * FROM semesters ORDER BY start_date DESC")
        semesters = cursor.fetchall()
        cursor.execute("SELECT * FROM courses ORDER BY title")
        courses = cursor.fetchall()
        cursor.execute("SELECT id, first_name, last_name FROM user WHERE role = 'teacher' ORDER BY last_name")
        instructors = cursor.fetchall()
        return render_template("admin/sections/list.html",
                            sections=sections,
                            semesters=semesters,
                            courses=courses,
                            instructors=instructors,
                            semester_id=semester_id,
                            course_id=course_id,
                            instructor_id=instructor_id,
                            pagination={
                                'page': page,
                                'per_page': per_page,
                                'total': total,
                                'pages': (total + per_page - 1) // per_page})
    return redirect(url_for('login'))

@app.route("/admin/sections/create", methods=['GET', 'POST'])
def create_section():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'GET':
            semester_id = request.args.get('semester_id', '')
            course_id = request.args.get('course_id', '')
            existing_sections = []
            next_section = 1
            if semester_id and course_id:
                try:
                    #get existing section numbers for course/semester
                    cursor.execute("""
                        SELECT DISTINCT section_number 
                        FROM sections 
                        WHERE semester_id = %s AND course_id = %s 
                        ORDER BY section_number
                        """, (semester_id, course_id))
                    existing_sections = [str(s['section_number']) for s in cursor.fetchall()]
                    #calculate next available section number
                    if existing_sections:
                        next_section = max(int(s) for s in existing_sections) + 1
                except Exception as e:
                    flash(f'Error fetching existing sections: {str(e)}', 'danger')
            # Get all options for dropdowns
            cursor.execute("SELECT * FROM semesters ORDER BY start_date DESC")
            semesters = cursor.fetchall()
            cursor.execute("SELECT * FROM courses ORDER BY title")
            courses = cursor.fetchall()
            cursor.execute("SELECT id, first_name, last_name FROM user WHERE role = 'teacher' ORDER BY last_name")
            instructors = cursor.fetchall()
            return render_template("admin/sections/add_section.html",
                                section=None,
                                semesters=semesters,
                                courses=courses,
                                instructors=instructors,
                                existing_sections=existing_sections,
                                next_section=next_section,
                                semester_id=semester_id,
                                course_id=course_id,
                                days=[],
                                meeting_times={})
        elif request.method == 'POST':
            #process form submission
            course_id = request.form['course_id']
            semester_id = request.form['semester_id']
            section_number = request.form['section_number']
            instructor_id = request.form.get('instructor_id') or None
            max_capacity = request.form['max_capacity']
            #process meeting times (from array-style form fields)
            meetings = []
            meetings_data = request.form.to_dict(flat=False).get('meetings[]', [])
            for i in range(len(meetings_data) // 3):
                meetings.append({
                    'day': meetings_data[i*3],
                    'start_time': meetings_data[i*3+1],
                    'end_time': meetings_data[i*3+2]})
            try:
                mysql.connection.begin()
                #check for instructor time conflicts
                if instructor_id:
                    for meeting in meetings:
                        cursor.execute("""
                            SELECT s.section_id, c.title as course_title, s.day, 
                                   TIME_FORMAT(s.time_start, '%%H:%%i') as time_start, 
                                   TIME_FORMAT(s.time_end, '%%H:%%i') as time_end
                            FROM sections s
                            JOIN courses c ON s.course_id = c.course_id
                            WHERE s.instructor_id = %s
                            AND s.semester_id = %s
                            AND s.day = %s
                            AND (
                                (s.time_start < %s AND s.time_end > %s) OR
                                (s.time_start < %s AND s.time_end > %s) OR
                                (s.time_start >= %s AND s.time_end <= %s)
                            )
                            """, (instructor_id, semester_id, meeting['day'],
                                  meeting['end_time'], meeting['start_time'],
                                  meeting['end_time'], meeting['start_time'],
                                  meeting['start_time'], meeting['end_time']))
                        conflicts = cursor.fetchall()
                        if conflicts:
                            conflict_msgs = [ f"{c['course_title']} on {c['day']} ({c['time_start']}-{c['time_end']})" for c in conflicts]
                            flash(f'Instructor has scheduling conflicts with: {", ".join(conflict_msgs)}','danger')
                            return redirect(request.url)
                #create section meetings
                for meeting in meetings:
                    cursor.execute("""
                        INSERT INTO sections 
                        (course_id, semester_id, section_number, instructor_id, max_capacity, 
                         day, time_start, time_end)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (course_id, semester_id, section_number, instructor_id, max_capacity,
                             meeting['day'], meeting['start_time'], meeting['end_time']))
                mysql.connection.commit()
                flash('Section created successfully!', 'success')
                return redirect(url_for('admin_sections'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating section: {str(e)}', 'danger')
                return redirect(request.url)
    return redirect(url_for('login'))

@app.route("/admin/sections/<int:section_id>/edit", methods=['GET', 'POST'])
def edit_section(section_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'GET':
            cursor.execute("""
                SELECT s.*, c.title as course_title 
                FROM sections s
                JOIN courses c ON s.course_id = c.course_id
                WHERE s.section_id = %s
                """, (section_id,))
            section = cursor.fetchone()
            if not section:
                flash('Section not found!', 'danger')
                return redirect(url_for('admin_sections'))
            #get dropdown options
            cursor.execute("SELECT * FROM semesters ORDER BY start_date DESC")
            semesters = cursor.fetchall()
            cursor.execute("SELECT * FROM courses ORDER BY title")
            courses = cursor.fetchall()
            cursor.execute("SELECT id, first_name, last_name FROM user WHERE role = 'teacher' ORDER BY last_name")
            instructors = cursor.fetchall()
            return render_template("admin/sections/form.html",
                                section=section,
                                semesters=semesters,
                                courses=courses,
                                instructors=instructors)
        elif request.method == 'POST':
            #process section form data only
            course_id = request.form['course_id']
            semester_id = request.form['semester_id']
            section_number = request.form['section_number']
            instructor_id = request.form.get('instructor_id') or None
            max_capacity = request.form['max_capacity']
            try:
                cursor.execute("""
                    UPDATE sections 
                    SET course_id = %s,
                        semester_id = %s,
                        section_number = %s,
                        instructor_id = %s,
                        max_capacity = %s
                    WHERE section_id = %s
                    """, (course_id, semester_id, section_number, instructor_id, max_capacity, section_id))
                mysql.connection.commit()
                flash('Section updated successfully!', 'success')
                return redirect(url_for('admin_sections'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating section: {str(e)}', 'danger')
                return redirect(url_for('edit_section', section_id=section_id))
    return redirect(url_for('login'))

@app.route("/admin/sections/<int:section_id>/delete", methods=['POST'])
def delete_section(section_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            cursor.execute("SELECT section_number, course_id, instructor_id, semester_id FROM sections WHERE section_id = %s",(section_id,))
            section = cursor.fetchone()
            if not section:
                flash('Section not found!', 'danger')
                return redirect(url_for('admin_sections'))
            #check if section has enrollments
            cursor.execute(
                """SELECT COUNT(*) as enrollment_count 
                FROM enrollments e
                JOIN sections s ON e.section_id = s.section_id
                WHERE s.section_number = %s 
                AND s.course_id = %s 
                AND s.instructor_id <=> %s 
                AND s.semester_id = %s
                AND e.status = 'registered'""",
                (section['section_number'], section['course_id'], 
                 section['instructor_id'], section['semester_id']))
            if cursor.fetchone()['enrollment_count'] > 0:
                flash('Cannot delete section with active enrollments!', 'danger')
                return redirect(url_for('admin_sections'))
            #delete all sections in this group
            cursor.execute(
                """DELETE FROM sections 
                WHERE section_number = %s 
                AND course_id = %s 
                AND instructor_id <=> %s 
                AND semester_id = %s""",
                (section['section_number'], section['course_id'], 
                 section['instructor_id'], section['semester_id']))
            mysql.connection.commit()
            flash('Section deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting section: {str(e)}', 'danger')
        return redirect(url_for('admin_sections'))
    return redirect(url_for('login'))

#section schedule view
@app.route("/admin/schedule")
def admin_schedule():
    if 'loggedin' in session and session['role'] == 'admin':
        semester_id = request.args.get('semester_id', '')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get selected semester or current semester
        if not semester_id:
            cursor.execute("SELECT * FROM semesters WHERE is_current = TRUE LIMIT 1")
            current_semester = cursor.fetchone()
            if current_semester:
                semester_id = current_semester['semester_id']
        else:
            cursor.execute("SELECT * FROM semesters WHERE semester_id = %s", (semester_id,))
            current_semester = cursor.fetchone()
        #get all sections for the semester
        sections = []
        if semester_id:
            cursor.execute("""
                SELECT 
                    c.title,
                    s.day,
                    s.time_start,
                    s.time_end,
                    s.section_number,
                    CONCAT(u.first_name, ' ', u.last_name) as instructor,
                    s.section_id,
                    c.course_id
                FROM sections s
                JOIN courses c ON s.course_id = c.course_id
                LEFT JOIN user u ON s.instructor_id = u.id
                WHERE s.semester_id = %s
                ORDER BY s.day, s.time_start
            """, (semester_id,))
            sections = cursor.fetchall()
        #get semesters for dropdown
        cursor.execute("SELECT * FROM semesters ORDER BY start_date DESC")
        semesters = cursor.fetchall()
        #create schedule grid(8-6)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        hours = list(range(8, 19))#8-6
        schedule = {day: {hour: None for hour in hours} for day in days}
        
        for section in sections:
            day = section['day'].capitalize()
            try:
                #convert time format from timedelta(if needed) or string
                if isinstance(section['time_start'], str):
                    start_time = datetime.strptime(section['time_start'], '%H:%M:%S').time()
                else:
                    start_time = (datetime.min + section['time_start']).time()
                if isinstance(section['time_end'], str):
                    end_time = datetime.strptime(section['time_end'], '%H:%M:%S').time()
                else:
                    end_time = (datetime.min + section['time_end']).time()
                #format the time display
                formatted_start = start_time.strftime('%I:%M %p').lstrip('0')
                formatted_end = end_time.strftime('%I:%M %p').lstrip('0')
                course_time = f"{formatted_start} - {formatted_end}"
                start_hour = start_time.hour
                end_hour = end_time.hour
                duration_hours = end_hour - start_hour
                #mark all affected hours
                for hour_slot in range(start_hour, end_hour):
                    if hour_slot in schedule[day]:
                        schedule[day][hour_slot] = {
                            'title': section['title'],
                            'section': section['section_number'],
                            'instructor': section['instructor'],
                            'time': course_time,
                            'is_first': (hour_slot == start_hour),
                            'duration': duration_hours,
                            'is_continuation': (hour_slot > start_hour),
                            'section_id': section['section_id'],
                            'course_id': section['course_id']}
            except Exception as e:
                print(f"Error processing schedule for section {section['section_id']}: {e}")
                continue
        return render_template("admin/sections/schedule.html",
                            schedule=schedule,
                            days=days,
                            hours=hours,
                            semesters=semesters,
                            current_semester=current_semester)
    return redirect(url_for('login'))

# Financial Management Routes
@app.route("/admin/finances")
def admin_finances():
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    semester_id = request.args.get('semester_id', 'all')
    status = request.args.get('status', 'all')
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    finance_query = """
        SELECT fr.*, s.name as semester_name,
               CONCAT(u.first_name, ' ', u.last_name) as student_name,
               st.student_id
        FROM financial_record fr
        JOIN semesters s ON fr.semester_id = s.semester_id
        JOIN students st ON fr.student_id = st.student_id
        JOIN user u ON st.user_id = u.id
        WHERE 1=1"""
    count_query = """
        SELECT COUNT(*) as total 
        FROM financial_record fr
        JOIN students st ON fr.student_id = st.student_id
        JOIN user u ON st.user_id = u.id
        WHERE 1=1"""
    params = []
    if semester_id != 'all':
        finance_query += " AND fr.semester_id = %s"
        count_query += " AND fr.semester_id = %s"
        params.append(semester_id)
    if status != 'all':
        finance_query += " AND fr.status = %s"
        count_query += " AND fr.status = %s"
        params.append(status)
    if search:
        search = search.strip()
        search_term = f"%{search}%"
        try:
            student_id_search = int(search)
        except ValueError:
            student_id_search = 0
        finance_query += """
            AND (st.student_id = %s OR
                u.first_name LIKE %s OR
                u.last_name LIKE %s OR
                CONCAT(u.first_name, ' ', u.last_name) LIKE %s)"""
        count_query += """
            AND (st.student_id = %s OR
                u.first_name LIKE %s OR
                u.last_name LIKE %s OR
                CONCAT(u.first_name, ' ', u.last_name) LIKE %s)"""
        params.extend([student_id_search, search_term, search_term, search_term])    
    #get total count
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute(count_query, params)
    total = cursor.fetchone()['total']
    #get paginated results
    finance_query += " ORDER BY fr.due_date DESC LIMIT %s OFFSET %s"
    params.extend([per_page, (page-1)*per_page])
    cursor.execute(finance_query, params)
    records = cursor.fetchall()
    #get semesters for dropdown
    cursor.execute("SELECT * FROM semesters ORDER BY start_date DESC")
    semesters = cursor.fetchall()
    #calculate summary statistics
    cursor.execute("""
        SELECT 
            SUM(CASE WHEN status = 'paid' THEN amount ELSE 0 END) as paid_amount,
            SUM(CASE WHEN status = 'pending' THEN amount ELSE 0 END) as pending_amount,
            SUM(CASE WHEN status = 'overdue' THEN amount ELSE 0 END) as overdue_amount,
            COUNT(*) as total_records FROM financial_record""")
    summary = cursor.fetchone()
    return render_template("admin/finances/list.html",
                        records=records,
                        semesters=semesters,
                        semester_id=semester_id,
                        status=status,
                        search=search,
                        summary=summary,
                        pagination={
                            'page': page,
                            'per_page': per_page,
                            'total': total,
                            'pages': (total + per_page - 1) // per_page})
    
@app.route("/admin/finances/create", methods=['GET', 'POST'])
def create_financial_record():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            student_id = request.form['student_id']
            semester_id = request.form['semester_id']
            amount = request.form['amount']
            due_date = request.form['due_date']
            payment_method = request.form.get('payment_method', '')
            status = request.form['status']
            try:
                cursor.execute(
                    """INSERT INTO financial_record 
                    (student_id, semester_id, amount, due_date, payment_method, status)
                    VALUES (%s, %s, %s, %s, %s, %s)""",
                    (student_id, semester_id, amount, due_date, payment_method, status))
                mysql.connection.commit()
                flash('Financial record created successfully!', 'success')
                return redirect(url_for('admin_finances'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating financial record: {str(e)}', 'danger')
        #get options for dropdowns
        cursor.execute("SELECT * FROM semesters ORDER BY start_date DESC")
        semesters = cursor.fetchall()   
        cursor.execute("""
            SELECT s.student_id, u.first_name, u.last_name 
            FROM students s
            JOIN user u ON s.user_id = u.id
            ORDER BY u.last_name""")
        students = cursor.fetchall()
        return render_template("admin/finances/form.html", 
                            record=None, 
                            semesters=semesters,
                            students=students)
    return redirect(url_for('login'))

@app.route("/admin/finances/<int:record_id>/edit", methods=['GET', 'POST'])
def edit_financial_record(record_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            student_id = request.form['student_id']
            semester_id = request.form['semester_id']
            amount = request.form['amount']
            due_date = request.form['due_date']
            payment_method = request.form.get('payment_method', '')
            status = request.form['status'] 
            try:
                cursor.execute(
                    """UPDATE financial_record 
                    SET student_id=%s, semester_id=%s, amount=%s, due_date=%s, 
                        payment_method=%s, status=%s
                    WHERE record_id=%s""",
                    (student_id, semester_id, amount, due_date, 
                     payment_method, status, record_id) )
                mysql.connection.commit()
                flash('Financial record updated successfully!', 'success')
                return redirect(url_for('admin_finances'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating financial record: {str(e)}', 'danger')
        #get record data
        cursor.execute("""
            SELECT fr.*, s.name as semester_name,
                   CONCAT(u.first_name, ' ', u.last_name) as student_name
            FROM financial_record fr
            JOIN semesters s ON fr.semester_id = s.semester_id
            JOIN students st ON fr.student_id = st.student_id
            JOIN user u ON st.user_id = u.id
            WHERE fr.record_id = %s
        """, (record_id,))
        record = cursor.fetchone()
        if not record:
            flash('Financial record not found!', 'danger')
            return redirect(url_for('admin_finances'))
        #get options for dropdowns
        cursor.execute("SELECT * FROM semesters ORDER BY start_date DESC")
        semesters = cursor.fetchall()
        cursor.execute("""
            SELECT s.student_id, u.first_name, u.last_name 
            FROM students s
            JOIN user u ON s.user_id = u.id
            ORDER BY u.last_name """)
        students = cursor.fetchall()
        return render_template("admin/finances/form.html", 
                            record=record, 
                            semesters=semesters,
                            students=students)
    return redirect(url_for('login'))

@app.route("/admin/finances/<int:record_id>/delete", methods=['POST'])
def delete_financial_record(record_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            cursor.execute("DELETE FROM financial_record WHERE record_id = %s", (record_id,))
            mysql.connection.commit()
            flash('Financial record deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting financial record: {str(e)}', 'danger')
        return redirect(url_for('admin_finances'))
    return redirect(url_for('login'))

@app.route("/admin/finances/reports")
def financial_reports():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #semester-wise summary
        cursor.execute("""
            SELECT s.semester_id, s.name, 
                   SUM(fr.amount) as total_amount,
                   SUM(CASE WHEN fr.status = 'paid' THEN fr.amount ELSE 0 END) as paid_amount,
                   SUM(CASE WHEN fr.status = 'pending' THEN fr.amount ELSE 0 END) as pending_amount,
                   SUM(CASE WHEN fr.status = 'overdue' THEN fr.amount ELSE 0 END) as overdue_amount
            FROM financial_record fr
            JOIN semesters s ON fr.semester_id = s.semester_id
            GROUP BY s.semester_id
            ORDER BY s.start_date DESC""")
        semester_summary = cursor.fetchall()
        #recent transactions
        cursor.execute("""
            SELECT fr.*, s.name as semester_name,
                   CONCAT(u.first_name, ' ', u.last_name) as student_name
            FROM financial_record fr
            JOIN semesters s ON fr.semester_id = s.semester_id
            JOIN students st ON fr.student_id = st.student_id
            JOIN user u ON st.user_id = u.id
            ORDER BY fr.due_date DESC
            LIMIT 10""")
        recent_transactions = cursor.fetchall()
        #payment methods breakdown
        cursor.execute("""
            SELECT payment_method, 
                   COUNT(*) as transaction_count,
                   SUM(amount) as total_amount
            FROM financial_record
            WHERE payment_method != ''
            GROUP BY payment_method
            ORDER BY total_amount DESC""")
        payment_methods = cursor.fetchall()
        return render_template("admin/finances/reports.html",
                            semester_summary=semester_summary,
                            recent_transactions=recent_transactions,
                            payment_methods=payment_methods)
    return redirect(url_for('login'))

@app.route("/admin/finances/export")
def export_financial_data():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get all financial records
        cursor.execute("""
            SELECT fr.*, s.name as semester_name,
                   CONCAT(u.first_name, ' ', u.last_name) as student_name
            FROM financial_record fr
            JOIN semesters s ON fr.semester_id = s.semester_id
            JOIN students st ON fr.student_id = st.student_id
            JOIN user u ON st.user_id = u.id
            ORDER BY fr.due_date DESC""")
        records = cursor.fetchall()
        #create CSV output
        output = io.StringIO()
        writer = csv.writer(output)
        #write header
        writer.writerow(['Record ID', 'Student', 'Semester', 'Amount', 'Due Date', 'Payment Method', 'Status'])
        #write data
        for record in records:
            writer.writerow([
                record['record_id'],
                record['student_name'],
                record['semester_name'],
                record['amount'],
                record['due_date'].strftime('%Y-%m-%d') if record['due_date'] else '',
                record['payment_method'],
                record['status']])
        #prepare response
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=financial_records.csv'
        response.headers['Content-type'] = 'text/csv'
        return response
    return redirect(url_for('login'))

#library management routes
@app.route("/admin/library/books")
def admin_library_books():
    if 'loggedin' in session and session['role'] == 'admin':
        search = request.args.get('search', '')
        category = request.args.get('category', 'all')
        availability = request.args.get('availability', 'all')
        page = request.args.get('page', 1, type=int)
        per_page = 10
        book_query = "SELECT * FROM books WHERE 1=1"
        count_query = "SELECT COUNT(*) as total FROM books WHERE 1=1"
        params = []
        if search:
            book_query += " AND (title LIKE %s OR author LIKE %s OR isbn = %s)"
            count_query += " AND (title LIKE %s OR author LIKE %s OR isbn = %s)"
            params.extend([f"%{search}%", f"%{search}%", search]) 
        if category != 'all':
            book_query += " AND category = %s"
            count_query += " AND category = %s"
            params.append(category)
        if availability != 'all':
            if availability == 'available':
                book_query += " AND available_copies > 0"
                count_query += " AND available_copies > 0"
            else:
                book_query += " AND available_copies = 0"
                count_query += " AND available_copies = 0"      
        #get total count
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        #get paginated results
        book_query += " ORDER BY title LIMIT %s OFFSET %s"
        params.extend([per_page, (page-1)*per_page])
        cursor.execute(book_query, params)
        books = cursor.fetchall()
        #get distinct categories for filter
        cursor.execute("SELECT DISTINCT category FROM books WHERE category IS NOT NULL AND category != '' ORDER BY category")
        categories = [row['category'] for row in cursor.fetchall()]
        return render_template("admin/library/books/list.html",
                            books=books,
                            categories=categories,
                            search=search,
                            category=category,
                            availability=availability,
                            pagination={
                                'page': page,
                                'per_page': per_page,
                                'total': total,
                                'pages': (total + per_page - 1) // per_page})
    return redirect(url_for('login'))

@app.route("/admin/library/books/create", methods=['GET', 'POST'])
def create_book():
    if 'loggedin' in session and session['role'] == 'admin':
        if request.method == 'POST':
            isbn = request.form.get('isbn', '')
            title = request.form['title']
            author = request.form['author']
            publisher = request.form.get('publisher', '')
            year = request.form.get('year', '')
            category = request.form.get('category', '')
            total_copies = request.form['total_copies']
            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            try:
                cursor.execute(
                    """INSERT INTO books 
                    (isbn, title, author, publisher, year, category, total_copies, available_copies)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                    (isbn, title, author, publisher, year, category, total_copies, total_copies))
                mysql.connection.commit()
                flash('Book added successfully!', 'success')
                return redirect(url_for('admin_library_books'))
            except MySQLdb.IntegrityError as e:
                mysql.connection.rollback()
                flash('ISBN already exists in the system!', 'danger')
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error adding book: {str(e)}', 'danger')
        return render_template("admin/library/books/form.html", book=None)
    return redirect(url_for('login'))

@app.route("/admin/library/books/<int:book_id>/edit", methods=['GET', 'POST'])
def edit_book(book_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            isbn = request.form.get('isbn', '')
            title = request.form['title']
            author = request.form['author']
            publisher = request.form.get('publisher', '')
            year = request.form.get('year', '')
            category = request.form.get('category', '')
            total_copies = request.form['total_copies']
            try: #cutrrent available copies
                cursor.execute("SELECT available_copies, total_copies FROM books WHERE book_id = %s", (book_id,))
                current = cursor.fetchone()
                #new available copies
                copies_diff = int(total_copies) - current['total_copies']
                new_available = current['available_copies'] + copies_diff
                cursor.execute(
                    """UPDATE books 
                    SET isbn=%s, title=%s, author=%s, publisher=%s, 
                        year=%s, category=%s, total_copies=%s, available_copies=%s
                    WHERE book_id=%s""",
                    (isbn, title, author, publisher, year, category, total_copies, new_available, book_id))
                #handle file upload if present
                if 'cover_image' in request.files and request.files['cover_image'].filename:
                    file = request.files['cover_image']
                mysql.connection.commit()
                flash('Book updated successfully!', 'success')
                return redirect(url_for('admin_library_books'))
            except MySQLdb.IntegrityError as e:
                mysql.connection.rollback()
                flash('ISBN already exists in the system!', 'danger')
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating book: {str(e)}', 'danger')
        cursor.execute("SELECT * FROM books WHERE book_id = %s", (book_id,))
        book = cursor.fetchone()
        if not book:
            flash('Book not found!', 'danger')
            return redirect(url_for('admin_library_books'))
        return render_template("admin/library/books/form.html", book=book)
    return redirect(url_for('login'))

@app.route("/admin/library/books/<int:book_id>/delete", methods=['POST'])
def delete_book(book_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            #check if book has active loans
            cursor.execute("""
                SELECT COUNT(*) as loan_count 
                FROM book_reserve 
                WHERE book_id = %s AND status = 'active'
            """, (book_id,))
            if cursor.fetchone()['loan_count'] > 0:
                flash('Cannot delete book with active loans!', 'danger')
                return redirect(url_for('admin_library_books'))
            #delete book
            cursor.execute("DELETE FROM books WHERE book_id = %s", (book_id,))
            mysql.connection.commit()
            flash('Book deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting book: {str(e)}', 'danger')
        return redirect(url_for('admin_library_books'))
    return redirect(url_for('login'))

#loan management routes
@app.route("/admin/library/loans")
def admin_library_loans():
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    try:
        status = request.args.get('status', 'active')
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        per_page = 10
        base_query = """
            FROM book_reserve br
            JOIN books b ON br.book_id = b.book_id
            JOIN students s ON br.student_id = s.student_id
            JOIN user u ON s.user_id = u.id
            WHERE 1=1 """
        loan_query = """
            SELECT 
                br.*, 
                b.title as book_title, 
                b.isbn, 
                CONCAT(u.first_name, ' ', u.last_name) as student_name, 
                u.email as student_email,
                br.due_date,
                br.return_date,
                br.status
            """ + base_query
        count_query = "SELECT COUNT(*) as total " + base_query
        params = []
        if status != 'all':
            loan_query += " AND br.status = %s"
            count_query += " AND br.status = %s"
            params.append(status)
        if search:
            loan_query += """
                AND (b.title LIKE %s 
                OR u.first_name LIKE %s 
                OR u.last_name LIKE %s 
                OR b.isbn = %s)"""
            count_query += """
                AND (b.title LIKE %s 
                OR u.first_name LIKE %s 
                OR u.last_name LIKE %s 
                OR b.isbn = %s)"""
            search_term = f"%{search}%"
            params.extend([search_term, search_term, search_term, search])
        #get total count
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        #get paginated results
        loan_query += " ORDER BY br.due_date LIMIT %s OFFSET %s"
        params.extend([per_page, (page-1)*per_page])
        cursor.execute(loan_query, params)
        loans = cursor.fetchall()
        #data formatted for front-end
        for loan in loans:
            if isinstance(loan['due_date'], str):
                loan['due_date'] = datetime.strptime(loan['due_date'], '%Y-%m-%d').date()
            if loan['return_date'] and isinstance(loan['return_date'], str):
                loan['return_date'] = datetime.strptime(loan['return_date'], '%Y-%m-%d').date()
        #calculate overdue loans
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM book_reserve 
            WHERE status = 'active' AND due_date < CURDATE()""")
        overdue_count = cursor.fetchone()['count']
        return render_template("admin/library/loans/list.html",
                            loans=loans,
                            status=status,
                            search=search,
                            overdue_count=overdue_count,
                            datetime=datetime,
                            pagination={
                                'page': page,
                                'per_page': per_page,
                                'total': total,
                                'pages': (total + per_page - 1) // per_page})
    except Exception as e:
        app.logger.error(f"Error in admin_library_loans: {str(e)}")
        flash("An error occurred while loading loans", "danger")
        return redirect(url_for('admin_dashboard'))

@app.route("/admin/library/loans/create", methods=['GET', 'POST'])
def create_loan():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            book_id = request.form['book_id']
            student_id = request.form['student_id']
            due_date = request.form['due_date']
            type = 'borrow'
            try:
                mysql.connection.begin()
                #check book availability
                cursor.execute("""
                    SELECT available_copies 
                    FROM books 
                    WHERE book_id = %s FOR UPDATE
                """, (book_id,))
                book = cursor.fetchone()
                if not book or book['available_copies'] <= 0:
                    flash('Book is not available for loan!', 'danger')
                    return redirect(url_for('admin_library_loans'))
                #create loan record
                cursor.execute(
                    """INSERT INTO book_reserve 
                    (book_id, student_id, type, due_date, status)
                    VALUES (%s, %s, %s, %s, 'active')""",
                    (book_id, student_id, type, due_date))
                #update available copies
                cursor.execute("UPDATE books SET available_copies = available_copies - 1 WHERE book_id = %s",(book_id,))
                mysql.connection.commit()
                flash('Loan created successfully!', 'success')
                return redirect(url_for('admin_library_loans'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating loan: {str(e)}', 'danger')
        #get available books
        cursor.execute("""
            SELECT b.* 
            FROM books b
            WHERE b.available_copies > 0
            ORDER BY b.title""")
        available_books = cursor.fetchall()
        #get active students
        cursor.execute("""
            SELECT s.student_id, u.first_name, u.last_name 
            FROM students s
            JOIN user u ON s.user_id = u.id
            WHERE u.is_active = TRUE
            ORDER BY u.last_name""")
        students = cursor.fetchall()
        #default due date(2 weeks from now)
        default_due_date = (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d')
        return render_template("admin/library/loans/form.html",
                            available_books=available_books,
                            students=students,
                            default_due_date=default_due_date)
    return redirect(url_for('login'))

@app.route("/admin/library/loans/<int:reserve_id>/return", methods=['POST'])
def return_loan(reserve_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            mysql.connection.begin()
            #get loan details
            cursor.execute("""
                SELECT book_id, student_id 
                FROM book_reserve 
                WHERE reserve_id = %s AND status = 'active'
            """, (reserve_id,))
            loan = cursor.fetchone()
            if not loan:
                flash('Active loan not found!', 'danger')
                return redirect(url_for('admin_library_loans'))
            #update loan status
            cursor.execute("""
                UPDATE book_reserve 
                SET return_date = CURDATE(), status = 'returned' 
                WHERE reserve_id = %s
            """, (reserve_id,))
            #update available copies
            cursor.execute("""
                UPDATE books 
                SET available_copies = available_copies + 1 
                WHERE book_id = %s
            """, (loan['book_id'],))
            mysql.connection.commit()
            flash('Book returned successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error returning book: {str(e)}', 'danger')
        return redirect(url_for('admin_library_loans'))
    return redirect(url_for('login'))

@app.route("/admin/library/loans/<int:reserve_id>/renew", methods=['POST'])
def renew_loan(reserve_id):
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        #loan existence and is_active
        cursor.execute("""
            SELECT book_id, status, due_date 
            FROM book_reserve 
            WHERE reserve_id = %s
        """, (reserve_id,))
        loan = cursor.fetchone()
        if not loan:
            flash('Loan not found!', 'danger')
            return redirect(url_for('admin_library_loans'))
        if loan['status'] != 'active':
            flash('Only active loans can be renewed!', 'danger')
            return redirect(url_for('admin_library_loans'))
        #calculate new due date(2 weeks from today)
        new_due_date = datetime.now().date() + timedelta(days=14)
        #update the loan record
        cursor.execute("""
            UPDATE book_reserve 
            SET due_date = %s
            WHERE reserve_id = %s
        """, (new_due_date, reserve_id))
        mysql.connection.commit()
        flash(f'Loan renewed successfully! New due date: {new_due_date.strftime("%Y-%m-%d")}', 'success')
    except MySQLdb.Error as e:
        mysql.connection.rollback()
        app.logger.error(f"Database error renewing loan {reserve_id}: {str(e)}")
        flash('Database error while renewing loan. Please try again.', 'danger')
    except Exception as e:
        mysql.connection.rollback()
        app.logger.error(f"Unexpected error renewing loan {reserve_id}: {str(e)}")
        flash('Unexpected error renewing loan. Please try again.', 'danger')
    finally:
        cursor.close()
    return redirect(url_for('admin_library_loans'))

@app.route("/admin/library/reports")
def library_reports():
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #most popular books
        cursor.execute("""
            SELECT b.book_id, b.title, b.author, 
                   COUNT(br.reserve_id) as loan_count
            FROM book_reserve br
            JOIN books b ON br.book_id = b.book_id
            GROUP BY b.book_id
            ORDER BY loan_count DESC
            LIMIT 10""")
        popular_books = cursor.fetchall()
        #overdue books
        cursor.execute("""
            SELECT br.*, b.title, b.isbn,
                   CONCAT(u.first_name, ' ', u.last_name) as student_name,
                   DATEDIFF(CURDATE(), br.due_date) as days_overdue
            FROM book_reserve br
            JOIN books b ON br.book_id = b.book_id
            JOIN students s ON br.student_id = s.student_id
            JOIN user u ON s.user_id = u.id
            WHERE br.status = 'active' AND br.due_date < CURDATE()
            ORDER BY days_overdue DESC
            LIMIT 10""")
        overdue_books = cursor.fetchall()
        #loan statistics
        cursor.execute("""
            SELECT 
                COUNT(*) as total_loans,
                SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_loans,
                SUM(CASE WHEN status = 'returned' THEN 1 ELSE 0 END) as returned_loans,
                SUM(CASE WHEN status = 'overdue' THEN 1 ELSE 0 END) as overdue_loans
            FROM book_reserve""")
        loan_stats = cursor.fetchone()
        return render_template("admin/library/reports.html",
                            popular_books=popular_books,
                            overdue_books=overdue_books,
                            loan_stats=loan_stats)
    return redirect(url_for('login'))

@app.route("/admin/library/reports/export")
def export_library_reports():
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    #get all report data (not just top 10 like in the view)
    cursor.execute("""
        SELECT 
            b.title, b.author, b.isbn,
            CONCAT(u.first_name, ' ', u.last_name) as student_name,
            br.due_date,
            CASE 
                WHEN br.status = 'active' AND br.due_date < CURDATE() THEN 'overdue'
                ELSE br.status
            END as status,
            DATEDIFF(CURDATE(), br.due_date) as days_overdue
        FROM book_reserve br
        JOIN books b ON br.book_id = b.book_id
        LEFT JOIN students s ON br.student_id = s.student_id
        LEFT JOIN user u ON s.user_id = u.id
        ORDER BY days_overdue DESC""")
    data = cursor.fetchall()
    #create CSV output
    output = io.StringIO()
    writer = csv.writer(output)
    #write header
    writer.writerow(['Title', 'Author', 'ISBN', 'Student', 'Due Date', 'Status', 'Days Overdue'])
    #write data
    for row in data:
        writer.writerow([
            row['title'],
            row['author'],
            row['isbn'],
            row['student_name'],
            row['due_date'].strftime('%Y-%m-%d') if row['due_date'] else '',
            row['status'],
            row['days_overdue'] or ''])
    #prepare response
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=library_reports.csv'
    response.headers['Content-type'] = 'text/csv'
    return response

#internship management routes
@app.route("/admin/internships")
def admin_internships():
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    status_filter = request.args.get('status', 'all')
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    base_query = "SELECT * FROM internship WHERE 1=1"
    count_query = "SELECT COUNT(*) as total FROM internship WHERE 1=1"
    params = []
    if status_filter != 'all':
        is_active = status_filter == 'active'
        base_query += " AND is_active = %s"
        count_query += " AND is_active = %s"
        params.append(is_active)
    #search functionality
    if search:
        search = search.strip()
        search_term = f"%{search}%"
        base_query += """
            AND (title LIKE %s 
                 OR company LIKE %s 
                 OR location LIKE %s)"""
        count_query += """
            AND (title LIKE %s 
                 OR company LIKE %s 
                 OR location LIKE %s)"""
        params.extend([search_term, search_term, search_term])
    #get total count
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute(count_query, params)
    total = cursor.fetchone()['total']
    #get paginated results
    base_query += " ORDER BY posted_date DESC LIMIT %s OFFSET %s"
    params.extend([per_page, (page-1)*per_page])
    cursor.execute(base_query, params)
    internships = cursor.fetchall()
    return render_template("admin/internships/list.html",
                         internships=internships,
                         status_filter=status_filter,
                         search=search,
                         pagination={
                             'page': page,
                             'per_page': per_page,
                             'total': total,
                             'pages': (total + per_page - 1) // per_page})

@app.route("/admin/internships/create", methods=['GET', 'POST'])
def create_internship():
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            title = request.form['title']
            company = request.form['company']
            description = request.form['description']
            location = request.form['location']
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            application_deadline = request.form['application_deadline']
            link = request.form['link']
            is_active = 1 if 'is_active' in request.form else 0
            cursor = mysql.connection.cursor()
            cursor.execute("""
                INSERT INTO internship 
                (title, company, description, location, start_date, end_date, 
                 application_deadline, posted_date, is_active, link)
                VALUES (%s, %s, %s, %s, %s, %s, %s, CURDATE(), %s, %s)
            """, (title, company, description, location, start_date, end_date, 
                  application_deadline, is_active, link))
            mysql.connection.commit()
            flash('Internship created successfully!', 'success')
            return redirect(url_for('admin_internships'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error creating internship: {str(e)}', 'danger')
    return render_template("admin/internships/form.html", internship=None)

@app.route("/admin/internships/<int:inter_id>/edit", methods=['GET', 'POST'])
def edit_internship(inter_id):
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    if request.method == 'POST':
        try:
            title = request.form['title']
            company = request.form['company']
            description = request.form['description']
            location = request.form['location']
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            application_deadline = request.form['application_deadline']
            link = request.form['link']
            is_active = 1 if 'is_active' in request.form else 0
            cursor.execute("""
                UPDATE internship SET
                    title = %s,
                    company = %s,
                    description = %s,
                    location = %s,
                    start_date = %s,
                    end_date = %s,
                    application_deadline = %s,
                    is_active = %s,
                    link = %s
                WHERE inter_id = %s
            """, (title, company, description, location, start_date, end_date,
                  application_deadline, is_active, link, inter_id))
            mysql.connection.commit()
            flash('Internship updated successfully!', 'success')
            return redirect(url_for('admin_internships'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error updating internship: {str(e)}', 'danger')
    cursor.execute("SELECT * FROM internship WHERE inter_id = %s", (inter_id,))
    internship = cursor.fetchone()
    return render_template("admin/internships/form.html", internship=internship)

@app.route("/admin/internships/<int:inter_id>/delete", methods=['POST'])
def delete_internship(inter_id):
    if 'loggedin' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    try:
        cursor = mysql.connection.cursor()
        cursor.execute("DELETE FROM internship WHERE inter_id = %s", (inter_id,))
        mysql.connection.commit()
        flash('Internship deleted successfully!', 'success')
    except Exception as e:
        mysql.connection.rollback()
        flash(f'Error deleting internship: {str(e)}', 'danger')
    
    return redirect(url_for('admin_internships'))

#meetings listing
@app.route("/admin/sections/<int:section_id>/meetings")
def list_meetings(section_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get the parent section info
        cursor.execute("""
            SELECT s.*, c.title as course_title, 
                   CONCAT(u.first_name, ' ', u.last_name) as instructor_name
            FROM sections s
            JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN user u ON s.instructor_id = u.id
            WHERE s.section_id = %s
            """, (section_id,))
        parent_section = cursor.fetchone()
        if not parent_section:
            flash('Section not found!', 'danger')
            return redirect(url_for('admin_sections'))
        #get all meetings for this section group
        cursor.execute("""
            SELECT section_id, day, 
                   TIME_FORMAT(time_start, '%%H:%%i') as time_start, 
                   TIME_FORMAT(time_end, '%%H:%%i') as time_end
            FROM sections 
            WHERE course_id = %s 
              AND semester_id = %s 
              AND section_number = %s
            ORDER BY FIELD(day, 'monday', 'tuesday', 'wednesday', 
                          'thursday', 'friday', 'saturday', 'sunday'),
                     time_start
            """, (parent_section['course_id'], parent_section['semester_id'], parent_section['section_number']))
        meetings = cursor.fetchall()
        return render_template("admin/sections/list_meetings.html",
                            section=parent_section,
                            meetings=meetings)
    return redirect(url_for('login'))

@app.route("/admin/sections/<int:section_id>/meetings/add", methods=['GET', 'POST'])
def add_meeting(section_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        #get section info including instructor
        cursor.execute("""
            SELECT s.*, c.title as course_title, 
                   CONCAT(u.first_name, ' ', u.last_name) as instructor_name,
                   sem.name as semester_name
            FROM sections s
            JOIN courses c ON s.course_id = c.course_id
            JOIN semesters sem ON s.semester_id = sem.semester_id
            LEFT JOIN user u ON s.instructor_id = u.id
            WHERE s.section_id = %s
            """, (section_id,))
        section = cursor.fetchone()
        if not section:
            flash('Section not found!', 'danger')
            return redirect(url_for('admin_sections'))
        if request.method == 'POST':
            day = request.form['day']
            time_start = request.form['time_start']
            time_end = request.form['time_end']
            #validate whole hour times
            if not (time_start.endswith(':00') and time_end.endswith(':00')):
                flash('Meeting times must be on the hour (e.g., 9:00, 10:00)', 'danger')
                return render_template("admin/sections/add_meeting.html",
                                    section=section,
                                    day=day,
                                    time_start=time_start,
                                    time_end=time_end)
            try:
                mysql.connection.begin()
                #get all instructor's current schedule
                if section['instructor_id']:
                    cursor.execute("""
                        SELECT s.day, 
                               TIME_FORMAT(s.time_start, '%%H:%%i') as time_start, 
                               TIME_FORMAT(s.time_end, '%%H:%%i') as time_end,
                               c.title as course_title,
                               sem.name as semester_name
                        FROM sections s
                        JOIN courses c ON s.course_id = c.course_id
                        JOIN semesters sem ON s.semester_id = sem.semester_id
                        WHERE s.instructor_id = %s
                        AND s.day = %s
                        AND s.section_id != %s
                        ORDER BY s.time_start
                        """, (section['instructor_id'], day, section_id))
                    existing_meetings = cursor.fetchall()
                    #convert to time objects for comparison
                    new_start = datetime.strptime(time_start, '%H:%M').time()
                    new_end = datetime.strptime(time_end, '%H:%M').time()
                    conflicts = []
                    for meeting in existing_meetings:
                        existing_start = datetime.strptime(meeting['time_start'], '%H:%M').time()
                        existing_end = datetime.strptime(meeting['time_end'], '%H:%M').time()
                        #check for conflicts (same as student schedule logic)
                        if not (new_end <= existing_start or new_start >= existing_end):
                            conflicts.append({
                                'course': meeting['course_title'],
                                'semester': meeting['semester_name'],
                                'time': f"{meeting['time_start']}-{meeting['time_end']}"})
                    if conflicts:
                        conflict_msgs = []
                        for conflict in conflicts:
                            note = ""
                            if conflict['semester'] != section['semester_name']:
                                note = f" (Semester: {conflict['semester']})"
                            conflict_msgs.append(
                                f"{conflict['course']}{note} at {conflict['time']}")
                        flash('Instructor scheduling conflict detected with:<br>' + '<br>'.join(conflict_msgs),'danger')
                        return render_template("admin/sections/add_meeting.html",
                                           section=section,
                                           day=day,
                                           time_start=time_start,
                                           time_end=time_end,
                                           conflict_error=True)
                #insert new meeting if no conflicts
                cursor.execute("""
                    INSERT INTO sections 
                    (course_id, semester_id, section_number, instructor_id, max_capacity, 
                     day, time_start, time_end)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (section['course_id'], section['semester_id'], section['section_number'],
                         section['instructor_id'], section['max_capacity'],
                         day, time_start, time_end))
                mysql.connection.commit()
                flash('Meeting added successfully!', 'success')
                return redirect(url_for('list_meetings', section_id=section_id))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error adding meeting: {str(e)}', 'danger')
                return redirect(url_for('add_meeting', section_id=section_id))
        return render_template("admin/sections/add_meeting.html",section=section)
    return redirect(url_for('login'))

#edit meeting form
@app.route("/admin/sections/<int:meeting_id>/edit-meeting", methods=['GET', 'POST'])
def edit_meeting(meeting_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'GET':
            #get the meeting to edit
            cursor.execute("""
                SELECT s.*, c.title as course_title,
                       CONCAT(u.first_name, ' ', u.last_name) as instructor_name
                FROM sections s
                JOIN courses c ON s.course_id = c.course_id
                LEFT JOIN user u ON s.instructor_id = u.id
                WHERE s.section_id = %s
                """, (meeting_id,))
            meeting = cursor.fetchone()
            if not meeting:
                flash('Meeting not found!', 'danger')
                return redirect(url_for('admin_sections'))
            #format times for the form
            def format_db_time(time_obj):
                if not time_obj:
                    return ''
                if hasattr(time_obj, 'strftime'):
                    return time_obj.strftime('%H:%M')
                else:  # Assume it's a timedelta
                    total_seconds = time_obj.total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    return f"{hours:02d}:{minutes:02d}"
            meeting['time_start'] = format_db_time(meeting['time_start'])
            meeting['time_end'] = format_db_time(meeting['time_end'])
            return render_template("admin/sections/edit_meeting.html",meeting=meeting)
        elif request.method == 'POST':
            day = request.form['day']
            time_start = request.form['time_start']
            time_end = request.form['time_end']
            try:
                mysql.connection.begin()
                cursor.execute("""
                    SELECT s.*, c.title as course_title,
                           CONCAT(u.first_name, ' ', u.last_name) as instructor_name
                    FROM sections s
                    JOIN courses c ON s.course_id = c.course_id
                    LEFT JOIN user u ON s.instructor_id = u.id
                    WHERE s.section_id = %s
                    """, (meeting_id,))
                meeting = cursor.fetchone()
                if not meeting:
                    flash('Meeting not found!', 'danger')
                    return redirect(url_for('admin_sections'))
                # Check for instructor time conflicts
                if meeting['instructor_id']:
                    cursor.execute("""
                        SELECT s.section_id, c.title as course_title, s.day, 
                               TIME_FORMAT(s.time_start, '%%H:%%i') as time_start, 
                               TIME_FORMAT(s.time_end, '%%H:%%i') as time_end
                        FROM sections s
                        JOIN courses c ON s.course_id = c.course_id
                        WHERE s.instructor_id = %s
                        AND s.semester_id = %s
                        AND s.day = %s
                        AND s.section_id != %s
                        AND (
                            (s.time_start < %s AND s.time_end > %s) OR  -- New meeting starts during existing
                            (s.time_start < %s AND s.time_end > %s) OR  -- New meeting ends during existing
                            (s.time_start >= %s AND s.time_end <= %s)    -- New meeting completely overlaps
                        )
                        """, (meeting['instructor_id'], meeting['semester_id'], day,
                              meeting_id,
                              time_end, time_start,
                              time_end, time_start,
                              time_start, time_end))
                    conflicts = cursor.fetchall()
                    if conflicts:
                        conflict_msgs = [f"{c['course_title']} on {c['day']} ({c['time_start']}-{c['time_end']})"for c in conflicts]
                        flash(f'Instructor has scheduling conflicts with: {", ".join(conflict_msgs)}', 'danger')
                        #preserve form inputs
                        meeting.update({'day': day,'time_start': time_start,'time_end': time_end})
                        return render_template("admin/sections/edit_meeting.html",meeting=meeting)
                #update the meeting
                cursor.execute("""
                    UPDATE sections 
                    SET day = %s,
                        time_start = %s,
                        time_end = %s
                    WHERE section_id = %s
                    """, (day, time_start, time_end, meeting_id))
                mysql.connection.commit()
                flash('Meeting updated successfully!', 'success')
                #find the parent section for redirect
                cursor.execute("""
                    SELECT section_id 
                    FROM sections 
                    WHERE course_id = %s 
                      AND semester_id = %s 
                      AND section_number = %s
                    ORDER BY section_id ASC  # The first one is likely the parent
                    LIMIT 1
                    """, (meeting['course_id'], meeting['semester_id'], meeting['section_number']))
                parent_section = cursor.fetchone()
                if not parent_section:
                    flash('Parent section not found!', 'danger')
                    return redirect(url_for('admin_sections'))
                return redirect(url_for('list_meetings', section_id=parent_section['section_id']))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating meeting: {str(e)}', 'danger')
                return redirect(url_for('edit_meeting', meeting_id=meeting_id))
    return redirect(url_for('login'))

@app.route("/admin/sections/<int:meeting_id>/delete-meeting", methods=['POST'])
def delete_meeting(meeting_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            mysql.connection.begin()
            #get the parent section info
            cursor.execute("""
                SELECT s.section_id, s.section_number, s.course_id, 
                       s.semester_id, s.instructor_id
                FROM sections s WHERE s.section_id = %s
                """, (meeting_id,))
            meeting = cursor.fetchone()
            if not meeting:
                flash('Meeting not found!', 'danger')
                return redirect(url_for('admin_sections'))
            #delete the meeting
            cursor.execute("DELETE FROM sections WHERE section_id = %s", (meeting_id,))
            #check if section still has any meetings
            cursor.execute("""
                SELECT section_id FROM sections 
                WHERE course_id = %s 
                AND section_number = %s 
                AND semester_id = %s
                AND (instructor_id = %s OR (instructor_id IS NULL AND %s IS NULL))
                """, (meeting['course_id'], meeting['section_number'], 
                     meeting['semester_id'], meeting['instructor_id'], meeting['instructor_id']))
            remaining_meetings = cursor.fetchall()
            mysql.connection.commit()
            if remaining_meetings:  #if section still has meetings
                #get one of the remaining section_ids to redirect to
                remaining_section_id = remaining_meetings[0]['section_id']
                flash('Meeting deleted successfully!', 'success')
                return redirect(url_for('list_meetings', section_id=remaining_section_id))
            else:  #if this was the last meeting of the section
                flash('Meeting deleted successfully! (No more meetings in this section)', 'success')
                return redirect(url_for('admin_sections'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting meeting: {str(e)}', 'danger')
            return redirect(url_for('admin_sections'))
    return redirect(url_for('login'))

#student management routes
@app.route("/admin/students")
def admin_students():
    if 'loggedin' in session and session['role'] == 'admin':
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        per_page = 10
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        query = """
            SELECT s.*, u.first_name, u.last_name, u.email, u.is_active 
            FROM students s
            JOIN user u ON s.user_id = u.id"""
        params = []
        if search:
            query += " WHERE u.first_name LIKE %s OR u.last_name LIKE %s OR u.email LIKE %s OR s.major LIKE %s OR s.student_id LIKE %s"
            params.extend([f"%{search}%"] * 5)
        #count total records for pagination
        count_query = "SELECT COUNT(*) as total FROM (" + query + ") as count_query"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        #add pagination
        query += " LIMIT %s OFFSET %s"
        params.extend([per_page, (page - 1) * per_page])
        cursor.execute(query, params)
        students = cursor.fetchall()
        pagination = {'page': page,'per_page': per_page,'total': total,'pages': (total + per_page - 1) // per_page}
        return render_template("admin/users/students/list.html", 
                             students=students, 
                             pagination=pagination,
                             search=search)
    return redirect(url_for('login'))

@app.route("/admin/students/create", methods=['GET', 'POST'])
def create_student():
    if 'loggedin' in session and session['role'] == 'admin':
        if request.method == 'POST':
            try:
                student_data = {
                    'first_name': request.form['first_name'],
                    'last_name': request.form['last_name'],
                    'email': request.form['email'],
                    'date_of_birth': request.form['date_of_birth'],
                    'gender': request.form.get('gender'),
                    'student_id': request.form['student_id'],
                    'major': request.form['major'],
                    'program': request.form['program'],
                    'enrollment_date': request.form['enrollment_date'],
                    'current_semester': request.form.get('current_semester'),
                    'address': request.form.get('address'),
                    'phone': request.form.get('phone'),
                    'is_active': 1 if request.form.get('is_active') else 0}
                cursor = mysql.connection.cursor()
                mysql.connection.begin()
                cursor.execute("""INSERT INTO user (first_name, last_name, email, role, is_active) VALUES (%s, %s, %s, 'student', %s)
                """, (student_data['first_name'],
                    student_data['last_name'],
                    student_data['email'],
                    student_data['is_active']))
                user_id = cursor.lastrowid
                #create student record
                cursor.execute("""
                    INSERT INTO students (
                        user_id, student_id, date_of_birth, gender, 
                        address, phone, major, program, 
                        enrollment_date, current_semester)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (user_id,
                    student_data['student_id'],
                    student_data['date_of_birth'],
                    student_data['gender'],
                    student_data['address'],
                    student_data['phone'],
                    student_data['major'],
                    student_data['program'],
                    student_data['enrollment_date'],
                    student_data['current_semester']))
                mysql.connection.commit()
                flash('Student created successfully!', 'success')
                return redirect(url_for('admin_students'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating student: {str(e)}', 'danger')
        return render_template("admin/users/students/form.html")
    return redirect(url_for('login'))

@app.route("/admin/students/edit/<int:student_id>", methods=['GET', 'POST'])
def edit_student(student_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            try:
                student_data = {
                    'first_name': request.form['first_name'],
                    'last_name': request.form['last_name'],
                    'email': request.form['email'],
                    'date_of_birth': request.form['date_of_birth'],
                    'gender': request.form.get('gender'),
                    'student_id': request.form['student_id'],
                    'major': request.form['major'],
                    'program': request.form['program'],
                    'enrollment_date': request.form['enrollment_date'],
                    'current_semester': request.form.get('current_semester'),
                    'address': request.form.get('address'),
                    'phone': request.form.get('phone'),
                    'is_active': 1 if request.form.get('is_active') else 0}
                mysql.connection.begin()
                #update user table
                cursor.execute("""
                    UPDATE user 
                    SET first_name = %s, 
                        last_name = %s, 
                        email = %s, 
                        is_active = %s WHERE id = (SELECT user_id FROM students WHERE student_id = %s)
                """, (student_data['first_name'],
                    student_data['last_name'],
                    student_data['email'],
                    student_data['is_active'],
                    student_id))
                #update students table
                cursor.execute("""
                    UPDATE students 
                    SET student_id = %s,
                        date_of_birth = %s,
                        gender = %s,
                        address = %s,
                        phone = %s,
                        major = %s,
                        program = %s,
                        enrollment_date = %s,
                        current_semester = %s WHERE student_id = %s
                """, (student_data['student_id'],
                    student_data['date_of_birth'],
                    student_data['gender'],
                    student_data['address'],
                    student_data['phone'],
                    student_data['major'],
                    student_data['program'],
                    student_data['enrollment_date'],
                    student_data['current_semester'], student_id))
                mysql.connection.commit()
                flash('Student updated successfully!', 'success')
                return redirect(url_for('admin_students'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating student: {str(e)}', 'danger')
        #load existing students data
        cursor.execute("""
            SELECT s.*, u.first_name, u.last_name, u.email, u.is_active 
            FROM students s
            JOIN user u ON s.user_id = u.id
            WHERE s.student_id = %s
        """, (student_id,))
        student = cursor.fetchone()
        if not student:
            flash('Student not found!', 'danger')
            return redirect(url_for('admin_students'))
        return render_template("admin/users/students/form.html", student=student)
    return redirect(url_for('login'))

@app.route("/admin/students/delete/<int:student_id>", methods=['POST'])
def delete_student(student_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            mysql.connection.begin()
            cursor.execute("SELECT user_id FROM students WHERE student_id = %s", (student_id,))
            student = cursor.fetchone()
            if not student:
                flash('Student not found!', 'danger')
                return redirect(url_for('admin_students'))
            cursor.execute("DELETE FROM students WHERE student_id = %s", (student_id,))
            cursor.execute("DELETE FROM user WHERE id = %s", (student['user_id'],))
            mysql.connection.commit()
            flash('Student deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting student: {str(e)}', 'danger')
        return redirect(url_for('admin_students'))
    return redirect(url_for('login'))

@app.route("/admin/students/deactivate/<int:student_id>", methods=['POST'])
def deactivate_student(student_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            cursor.execute("SELECT user_id FROM students WHERE student_id = %s", (student_id,))
            student = cursor.fetchone()
            if not student:
                flash('Student not found!', 'danger')
                return redirect(url_for('admin_students'))
            cursor.execute("UPDATE user SET is_active = 0 WHERE id = %s", (student['user_id'],)) 
            mysql.connection.commit()
            flash('Student deactivated successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deactivating student: {str(e)}', 'danger')
        return redirect(url_for('admin_students'))
    return redirect(url_for('login'))


@app.route("/admin/students/import-export")
def student_import_export():
    if 'loggedin' in session and session['role'] == 'admin':
        return render_template("admin/users/students/import_export.html")
    return redirect(url_for('login'))

@app.route("/admin/students/export", methods=['POST'])
def export_students():
    if 'loggedin' in session and session['role'] == 'admin':
        format = request.form.get('export_format', 'csv')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("""
            SELECT s.*, u.first_name, u.last_name, u.email, u.is_active
            FROM students s
            JOIN user u ON s.user_id = u.id """)
        students = cursor.fetchall()
        if format == 'csv':
            si = StringIO()
            cw = csv.writer(si)
            #write header
            if students:
                cw.writerow(students[0].keys())
            #write data
            for student in students:
                cw.writerow(student.values())
            output = BytesIO()
            output.write(si.getvalue().encode('utf-8'))
            output.seek(0)
            return send_file(output,mimetype='text/csv',as_attachment=True,download_name='students_export.csv')
        elif format == 'json':
            output = BytesIO()
            output.write(json.dumps(students, indent=2).encode('utf-8'))
            output.seek(0)
            return send_file(output,mimetype='application/json',as_attachment=True,download_name='students_export.json')
        elif format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            #write header
            if students:
                ws.append(list(students[0].keys()))
            #write data
            for student in students:
                ws.append(list(student.values()))
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name='students_export.xlsx')
    return redirect(url_for('login'))

@app.route("/admin/students/import", methods=['POST'])
def import_students():
    if 'loggedin' in session and session['role'] == 'admin':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(url_for('student_import_export'))
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('student_import_export'))     
        update_existing = request.form.get('update_existing') == 'on'
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            if file.filename.endswith('.csv'):                
                csv_file = TextIOWrapper(file.stream, encoding='utf-8-sig')
                reader = csv.DictReader(csv_file)
                for row in reader:
                    process_student_row(row, cursor, update_existing)
            elif file.filename.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    process_student_row(row_dict, cursor, update_existing)
            elif file.filename.endswith('.json'):
                data = json.load(file)
                for student in data:
                    process_student_row(student, cursor, update_existing)
            mysql.connection.commit()
            flash('Students imported successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error importing students: {str(e)}', 'danger')
        return redirect(url_for('student_import_export'))
    return redirect(url_for('login'))

def process_student_row(row, cursor, update_existing):
    #extract data from row with defaults
    email = row.get('email', '').strip().lower()
    first_name = row.get('first_name', '').strip()
    last_name = row.get('last_name', '').strip()
    password = row.get('password', 'defaultpassword')
    role = 'student'
    #student profile data
    date_of_birth = row.get('date_of_birth')
    gender = row.get('gender', '').lower()
    address = row.get('address')
    phone = row.get('phone')
    major = row.get('major')
    program = row.get('program', 'bachelors')
    enrollment_date = row.get('enrollment_date')
    current_semester = row.get('current_semester')
    #check if user exists
    cursor.execute("SELECT id FROM user WHERE email = %s", (email,))
    existing_user = cursor.fetchone()
    if existing_user and not update_existing:
        return  #skip if exists and not updating
    hashed_password=hash_password(password)
    if existing_user and update_existing:
        #update existing user
        user_id = existing_user['id']
        cursor.execute("""
            UPDATE user 
            SET first_name = %s, last_name = %s, password = %s
            WHERE id = %s
        """, (first_name, last_name, hashed_password, user_id))
        #update student profile
        cursor.execute("""
            UPDATE students 
            SET date_of_birth = %s, gender = %s, address = %s, 
                phone = %s, major = %s, program = %s,
                enrollment_date = %s, current_semester = %s
            WHERE user_id = %s
        """, (
            date_of_birth, gender, address, phone, major, program,
            enrollment_date, current_semester, user_id))
    else:
        #create new user
        cursor.execute("""
            INSERT INTO user 
            (email, password, role, first_name, last_name)
            VALUES (%s, %s, %s, %s, %s)
        """, (email, hashed_password, role, first_name, last_name))
        user_id = cursor.lastrowid
        #create student profile
        cursor.execute("""
            INSERT INTO students 
            (user_id, date_of_birth, gender, address, phone, 
             major, program, enrollment_date, current_semester)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            user_id, date_of_birth, gender, address, phone,
            major, program, enrollment_date, current_semester))
        
# Teacher management routes
@app.route("/admin/teachers")
def admin_teachers():
    if 'loggedin' in session and session['role'] == 'admin':
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        per_page = 10
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        query = """
            SELECT t.*, u.first_name, u.last_name, u.email, u.is_active 
            FROM teacher t
            JOIN user u ON t.user_id = u.id"""
        params = []
        if search:
            query += " WHERE u.first_name LIKE %s OR u.last_name LIKE %s OR u.email LIKE %s OR t.teaching_program LIKE %s OR t.teacher_id LIKE %s"
            params.extend([f"%{search}%"] * 5)
        # Count total records for pagination
        count_query = "SELECT COUNT(*) as total FROM (" + query + ") as count_query"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        # Add pagination
        query += " LIMIT %s OFFSET %s"
        params.extend([per_page, (page - 1) * per_page])
        cursor.execute(query, params)
        teachers = cursor.fetchall()
        pagination = {'page': page,'per_page': per_page,'total': total,'pages': (total + per_page - 1) // per_page}
        return render_template("admin/users/teachers/list.html", 
                             teachers=teachers, 
                             pagination=pagination,
                             search=search)
    return redirect(url_for('login'))

@app.route("/admin/teachers/create", methods=['GET', 'POST'])
def create_teacher():
    if 'loggedin' in session and session['role'] == 'admin':
        if request.method == 'POST':
            try:
                teacher_data = {
                    'first_name': request.form['first_name'],
                    'last_name': request.form['last_name'],
                    'email': request.form['email'],
                    'birth_date': request.form['birth_date'],
                    'gender': request.form.get('gender'),
                    'teacher_id': request.form['teacher_id'],
                    'salary': request.form['salary'],
                    'teaching_program': request.form['teaching_program'],
                    'enrollment_date': request.form['enrollment_date'],
                    'address': request.form.get('address'),
                    'phone': request.form.get('phone'),
                    'is_active': 1 if request.form.get('is_active') else 0}
                cursor = mysql.connection.cursor()
                mysql.connection.begin()
                cursor.execute("""INSERT INTO user (first_name, last_name, email, role, is_active) 
                                VALUES (%s, %s, %s, 'teacher', %s)
                """, (teacher_data['first_name'],
                    teacher_data['last_name'],
                    teacher_data['email'],
                    teacher_data['is_active']))
                user_id = cursor.lastrowid
                # Create teacher record
                cursor.execute("""
                    INSERT INTO teacher (
                        user_id, teacher_id, birth_date, gender, 
                        address, phone, salary, teaching_program, 
                        enrollment_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (user_id,
                    teacher_data['teacher_id'],
                    teacher_data['birth_date'],
                    teacher_data['gender'],
                    teacher_data['address'],
                    teacher_data['phone'],
                    teacher_data['salary'],
                    teacher_data['teaching_program'],
                    teacher_data['enrollment_date']))
                mysql.connection.commit()
                flash('Teacher created successfully!', 'success')
                return redirect(url_for('admin_teachers'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error creating teacher: {str(e)}', 'danger')
        return render_template("admin/users/teachers/form.html")
    return redirect(url_for('login'))

@app.route("/admin/teachers/edit/<int:teacher_id>", methods=['GET', 'POST'])
def edit_teacher(teacher_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        if request.method == 'POST':
            try:
                teacher_data = {
                    'first_name': request.form['first_name'],
                    'last_name': request.form['last_name'],
                    'email': request.form['email'],
                    'birth_date': request.form['birth_date'],
                    'gender': request.form.get('gender'),
                    'teacher_id': request.form['teacher_id'],
                    'salary': request.form['salary'],
                    'teaching_program': request.form['teaching_program'],
                    'enrollment_date': request.form['enrollment_date'],
                    'address': request.form.get('address'),
                    'phone': request.form.get('phone'),
                    'is_active': 1 if request.form.get('is_active') else 0}
                mysql.connection.begin()
                # Update user table
                cursor.execute("""
                    UPDATE user 
                    SET first_name = %s, 
                        last_name = %s, 
                        email = %s, 
                        is_active = %s 
                    WHERE id = (SELECT user_id FROM teacher WHERE teacher_id = %s)
                """, (teacher_data['first_name'],
                    teacher_data['last_name'],
                    teacher_data['email'],
                    teacher_data['is_active'],
                    teacher_id))
                # Update teachers table
                cursor.execute("""
                    UPDATE teacher
                    SET teacher_id = %s,
                        birth_date = %s,
                        gender = %s,
                        address = %s,
                        phone = %s,
                        salary = %s,
                        teaching_program = %s,
                        enrollment_date = %s
                    WHERE teacher_id = %s
                """, (teacher_data['teacher_id'],
                    teacher_data['birth_date'],
                    teacher_data['gender'],
                    teacher_data['address'],
                    teacher_data['phone'],
                    teacher_data['salary'],
                    teacher_data['teaching_program'],
                    teacher_data['enrollment_date'], 
                    teacher_id))
                mysql.connection.commit()
                flash('Teacher updated successfully!', 'success')
                return redirect(url_for('admin_teachers'))
            except Exception as e:
                mysql.connection.rollback()
                flash(f'Error updating teacher: {str(e)}', 'danger')
        # Load existing teacher data
        cursor.execute("""
            SELECT t.*, u.first_name, u.last_name, u.email, u.is_active 
            FROM teacher t
            JOIN user u ON t.user_id = u.id
            WHERE t.teacher_id = %s
        """, (teacher_id,))
        teacher = cursor.fetchone()
        if not teacher:
            flash('Teacher not found!', 'danger')
            return redirect(url_for('admin_teachers'))
        return render_template("admin/users/teachers/form.html", teacher=teacher)
    return redirect(url_for('login'))

@app.route("/admin/teachers/delete/<int:teacher_id>", methods=['POST'])
def delete_teacher(teacher_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            mysql.connection.begin()
            cursor.execute("SELECT user_id FROM teacher WHERE teacher_id = %s", (teacher_id,))
            teacher = cursor.fetchone()
            if not teacher:
                flash('Teacher not found!', 'danger')
                return redirect(url_for('admin_teachers'))
            cursor.execute("DELETE FROM teacher WHERE teacher_id = %s", (teacher_id,))
            cursor.execute("DELETE FROM user WHERE id = %s", (teacher['user_id'],))
            mysql.connection.commit()
            flash('Teacher deleted successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deleting teacher: {str(e)}', 'danger')
        return redirect(url_for('admin_teachers'))
    return redirect(url_for('login'))

@app.route("/admin/teachers/deactivate/<int:teacher_id>", methods=['POST'])
def deactivate_teacher(teacher_id):
    if 'loggedin' in session and session['role'] == 'admin':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            cursor.execute("SELECT user_id FROM teacher WHERE teacher_id = %s", (teacher_id,))
            teacher = cursor.fetchone()
            if not teacher:
                flash('Teacher not found!', 'danger')
                return redirect(url_for('admin_teachers'))
            cursor.execute("UPDATE user SET is_active = 0 WHERE id = %s", (teacher['user_id'],)) 
            mysql.connection.commit()
            flash('Teacher deactivated successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error deactivating teacher: {str(e)}', 'danger')
        return redirect(url_for('admin_teachers'))
    return redirect(url_for('login'))

@app.route("/admin/teachers/import-export")
def teacher_import_export():
    if 'loggedin' in session and session['role'] == 'admin':
        return render_template("admin/users/teachers/import_export.html")
    return redirect(url_for('login'))

@app.route("/admin/teachers/export", methods=['POST'])
def export_teachers():
    if 'loggedin' in session and session['role'] == 'admin':
        format = request.form.get('export_format', 'csv')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("""
            SELECT t.*, u.first_name, u.last_name, u.email, u.is_active
            FROM teacher t
            JOIN user u ON t.user_id = u.id """)
        teachers = cursor.fetchall()
        if format == 'csv':
            si = StringIO()
            cw = csv.writer(si)
            # Write header
            if teachers:
                cw.writerow(teachers[0].keys())
            # Write data
            for teacher in teachers:
                cw.writerow(teacher.values())
            output = BytesIO()
            output.write(si.getvalue().encode('utf-8'))
            output.seek(0)
            return send_file(output,mimetype='text/csv',as_attachment=True,download_name='teachers_export.csv')
        elif format == 'json':
            output = BytesIO()
            output.write(json.dumps(teachers, indent=2).encode('utf-8'))
            output.seek(0)
            return send_file(output,mimetype='application/json',as_attachment=True,download_name='teachers_export.json')
        elif format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            # Write header
            if teachers:
                ws.append(list(teachers[0].keys()))
            # Write data
            for teacher in teachers:
                ws.append(list(teacher.values()))
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name='teachers_export.xlsx')
    return redirect(url_for('login'))

@app.route("/admin/teachers/import", methods=['POST'])
def import_teachers():
    if 'loggedin' in session and session['role'] == 'admin':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(url_for('teacher_import_export'))
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('teacher_import_export'))     
        update_existing = request.form.get('update_existing') == 'on'
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            if file.filename.endswith('.csv'):                
                csv_file = TextIOWrapper(file.stream, encoding='utf-8-sig')
                reader = csv.DictReader(csv_file)
                for row in reader:
                    process_teacher_row(row, cursor, update_existing)
            elif file.filename.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    process_teacher_row(row_dict, cursor, update_existing)
            elif file.filename.endswith('.json'):
                data = json.load(file)
                for teacher in data:
                    process_teacher_row(teacher, cursor, update_existing)
            mysql.connection.commit()
            flash('Teachers imported successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error importing teachers: {str(e)}', 'danger')
        return redirect(url_for('teacher_import_export'))
    return redirect(url_for('login'))

def process_teacher_row(row, cursor, update_existing):
    # Extract data from row with defaults
    email = row.get('email', '').strip().lower()
    first_name = row.get('first_name', '').strip()
    last_name = row.get('last_name', '').strip()
    password = row.get('password', 'defaultpassword')
    role = 'teacher'
    # Teacher profile data
    birth_date = row.get('birth_date')
    gender = row.get('gender', '').lower()
    address = row.get('address')
    phone = row.get('phone')
    salary = row.get('salary', 0)
    teaching_program = row.get('teaching_program')
    enrollment_date = row.get('enrollment_date')
    # Check if user exists
    cursor.execute("SELECT id FROM user WHERE email = %s", (email,))
    existing_user = cursor.fetchone()
    if existing_user and not update_existing:
        return  # Skip if exists and not updating
    hashed_password=hash_password(password)
    if existing_user and update_existing:
        # Update existing user
        user_id = existing_user['id']
        cursor.execute("""
            UPDATE user 
            SET first_name = %s, last_name = %s, password = %s
            WHERE id = %s
        """, (first_name, last_name, hashed_password, user_id))
        # Update teacher profile
        cursor.execute("""
            UPDATE teacher 
            SET birth_date = %s, gender = %s, address = %s, 
                phone = %s, salary = %s, teaching_program = %s,
                enrollment_date = %s
            WHERE user_id = %s
        """, (
            birth_date, gender, address, phone, salary, teaching_program,
            enrollment_date, user_id))
    else:
        # Create new user
        cursor.execute("""
            INSERT INTO user 
            (email, password, role, first_name, last_name)
            VALUES (%s, %s, %s, %s, %s)
        """, (email, hashed_password, role, first_name, last_name))
        user_id = cursor.lastrowid
        # Create teacher profile
        cursor.execute("""
            INSERT INTO teacher
            (user_id, birth_date, gender, address, phone, 
             salary, teaching_program, enrollment_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            user_id, birth_date, gender, address, phone,
            salary, teaching_program, enrollment_date))
    
#teacher functions

@app.route("/teacher/dashboard")
def teacher_dashboard():
    if 'loggedin' in session and session['role'] == 'teacher':
        # Get teacher_id from the teacher table based on user_id
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (session['userid'],))
        teacher = cursor.fetchone()
        
        if not teacher:
            return "Teacher record not found!", 403
        
        teacher_id = teacher['teacher_id']

        # Get the current semester
        current_semester = get_current_semester()

        # Get courses taught by the teacher in the current semester
        cursor.execute("""
            SELECT
                c.title,
                s.section_id,
                sem.name 
            FROM course_section_teacher cst
            JOIN courses c ON c.course_id = cst.course_id
            JOIN sections s ON s.section_id = cst.section_id
            JOIN semesters sem ON s.semester_id = sem.semester_id
            WHERE cst.teacher_id = %s""", 
            (session['userid'],))
        current_courses = cursor.fetchall()
        
        # Get upcoming assignments for the teacher's sections
        cursor.execute("""
            SELECT a.title, a.due_date, c.title AS course_title, s.section_id
            FROM assignments a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            WHERE s.instructor_id = %s AND a.due_date > NOW()
            ORDER BY a.due_date ASC
            LIMIT 5
        """, (session['userid'],))  # Again using user_id
        upcoming_assignments = cursor.fetchall()

        # Get announcements made by the teacher
        cursor.execute("""
            SELECT title, message, created_at
            FROM announcements
            WHERE teacher_id = %s AND created_at > NOW() - INTERVAL 30 DAY
            ORDER BY created_at DESC
            LIMIT 5
        """, (teacher_id,))
        recent_announcements = cursor.fetchall()

        # Get office hours for the teacher
        cursor.execute("""
            SELECT c.title, o.day, o.start_time, o.end_time
            FROM office_hours o
            JOIN courses c ON o.course_id = c.course_id          
            WHERE teacher_id = %s
        """, (teacher_id,))
        office_hours = cursor.fetchall()

        cursor.close()

        return render_template("teacher/dashboard.html",
                             current_semester=current_semester,
                             current_courses=current_courses,
                             upcoming_assignments=upcoming_assignments,
                             recent_announcements=recent_announcements,
                             office_hours=office_hours)
    return redirect(url_for('login'))

# Teacher profile routes
@app.route("/teacher/profile")
def teacher_profile():
    if 'loggedin' in session and session['role'] == 'teacher':
        user_id = session['userid']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('SELECT * FROM user WHERE id = %s', (user_id,))
        user = cursor.fetchone()
        
        cursor.execute('SELECT * FROM teacher WHERE user_id = %s', (user_id,))
        teacher = cursor.fetchone()
        
        if teacher:
            user.update(teacher)  # Combine user and teacher data
        
        return render_template("teacher/profile.html", user=user)
    return redirect(url_for('login'))

@app.route("/upload_teacher_profile_pic", methods=['POST'])
def upload_teacher_profile_pic():
    if 'loggedin' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))
    
    if 'profile_pic' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('teacher_profile'))
    
    file = request.files['profile_pic']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('teacher_profile'))
    
    if file and allowed_file(file.filename):
        user_id = session['userid']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get current profile pic filename to delete it later
        cursor.execute('SELECT profile_pic FROM teacher WHERE user_id = %s', (user_id,))
        teacher = cursor.fetchone()
        old_filename = teacher['profile_pic'] if teacher and teacher['profile_pic'] else None
        
        # Generate secure filename
        filename = secure_filename(f"teacher_{user_id}_{file.filename}")
        
        # Create complete file path
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        try:
            # Save the new file
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            print(filename)
            # Update database
            cursor.execute('UPDATE teacher SET profile_pic = %s WHERE user_id = %s', (filename, user_id))
            mysql.connection.commit()
            
            # Delete old file if it exists
            if old_filename:
                old_file_path = os.path.join(app.config['UPLOAD_FOLDER'], old_filename)
                if os.path.exists(old_file_path):
                    try:
                        os.remove(old_file_path)
                    except Exception as e:
                        print(f"Error deleting old profile picture: {e}")
                
            flash('Profile picture updated successfully!', 'success')
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error updating profile picture: {str(e)}', 'danger')
    else:
        flash('Allowed file types are: png, jpg, jpeg, gif', 'danger')
    
    return redirect(url_for('teacher_profile'))

@app.route('/teacher/courses')
def teacher_courses():
    if 'loggedin' in session and session['role'] == 'teacher':
        user_id = session['userid']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # First, get teacher_id using user_id
        cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (user_id,))
        teacher = cursor.fetchone()
        if not teacher:
            print("Teacher record not found")
            return redirect(url_for("teacher_dashboard"))

        teacher_id = teacher['teacher_id']

        # Fetch assigned courses
        query = """
            SELECT
                c.title AS course_title,
                s.section_number,
                sem.semester_id,
                sem.name AS semester_name,
                s.day,
                s.time_start,
                s.time_end
            FROM course_section_teacher cst
            JOIN courses c ON c.course_id = cst.course_id
            JOIN sections s ON s.section_id = cst.section_id
            JOIN semesters sem ON s.semester_id = sem.semester_id
            WHERE cst.teacher_id = %s
        """

        cursor.execute(query, (teacher_id,))
        assigned_courses = cursor.fetchall()

        return render_template("teacher/courses.html", assigned_courses=assigned_courses)
    return redirect(url_for('login'))

@app.route('/teacher/attendance', methods=['GET', 'POST'])
def teacher_attendance():
    if 'loggedin' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    teacher_user_id = session['userid']
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Get teacher_id
    cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (teacher_user_id,))
    teacher = cursor.fetchone()
    if not teacher:
        return "Teacher profile not found."

    teacher_id = teacher['teacher_id']

    # Get assigned courses and sections
    cursor.execute("""
       SELECT c.course_id, c.title, s.section_number, sem.name AS semester_name, sem.start_date, sem.end_date
        FROM course_section_teacher cst
        JOIN courses c ON c.course_id = cst.course_id
        JOIN sections s ON s.section_id = cst.section_id
        JOIN semesters sem ON sem.semester_id = cst.semester_id
        WHERE cst.teacher_id = %s
    """, (teacher_id,))
    assignments = cursor.fetchall()

    students = []
    attendance_map = {}  # Initialize attendance_map here
    
    if request.method == 'POST':
        course_id = request.form['course']
        section_id = request.form['section']
        attendance_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()

        # Get students for selected course/section
        cursor.execute("""
            SELECT e.student_id, u.first_name, u.last_name
            FROM enrollments e
            JOIN students st ON st.student_id = e.student_id
            JOIN user u ON u.id = st.user_id
            WHERE e.section_id = %s AND e.course_id = %s
        """, (section_id, course_id))
        students = cursor.fetchall()
        
        # Get existing attendance records
        cursor.execute("""
            SELECT student_id, status
            FROM attendance
            WHERE section_id = %s AND course_id = %s AND date = %s
        """, (section_id, course_id, attendance_date))
        for row in cursor.fetchall():
            attendance_map[row['student_id']] = row['status']

        # Handle attendance submission
        if 'submit_attendance' in request.form:
            for student in students:
                status = request.form.get(f"status_{student['student_id']}")
                if status:
                    cursor.execute("""
                        INSERT INTO attendance (section_id, student_id, course_id, status, date, duration)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE status=%s, duration=%s
                    """, (section_id, student['student_id'], course_id, status, attendance_date, 1.0, status, 1.0))
            mysql.connection.commit()
            flash("Attendance submitted successfully.", "success")
            return redirect(url_for('teacher_attendance'))

    return render_template(
        'teacher/attendance.html',
        assignments=assignments,
        students=students,
        attendance_map=attendance_map  # Now always defined
    )

def update_transcript(cursor, student_id, course_id, semester_id, final_result):
    """Update transcript with grade conversion based on final result"""
    # Ensure final_result is capped at 100
    final_result = min(float(final_result), 100.0)
    
    # Convert final result to grade and points based on grading system
    if final_result >= 90: grade, points = 'A+', 4.00
    elif final_result >= 86: grade, points = 'A', 4.00
    elif final_result >= 82: grade, points = 'A-', 3.67
    elif final_result >= 78: grade, points = 'B+', 3.33
    elif final_result >= 74: grade, points = 'B', 3.00
    elif final_result >= 70: grade, points = 'B-', 2.67
    elif final_result >= 66: grade, points = 'C+', 2.33
    elif final_result >= 62: grade, points = 'C', 2.00
    elif final_result >= 58: grade, points = 'C-', 1.67
    elif final_result >= 54: grade, points = 'D+', 1.33
    elif final_result >= 50: grade, points = 'D', 1.00
    else: grade, points = 'F', 0.00
    
    # Default to 'core' type since it's not in courses table
    course_type = 'core'
    
    # Update transcript
    cursor.execute("""
        INSERT INTO transcript (student_id, course_id, semester_id, grade, points, type)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            grade = VALUES(grade),
            points = VALUES(points),
            type = VALUES(type)
    """, (student_id, course_id, semester_id, grade, points, course_type))

@app.route('/teacher/grades', methods=['GET', 'POST'])
def teacher_grades():
    if 'loggedin' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (session['userid'],))
    teacher_row = cursor.fetchone()

    if not teacher_row:
        flash("Teacher not found for the logged-in user.", "danger")
        return redirect(url_for('dashboard'))

    teacher_id = teacher_row['teacher_id']
    cursor.execute("""
        SELECT c.course_id, c.title, s.section_number, s.semester_id, s.section_id
        FROM course_section_teacher cst
        JOIN courses c ON c.course_id = cst.course_id
        JOIN sections s ON s.section_id = cst.section_id
        WHERE cst.teacher_id = %s
    """, (teacher_id,))
    assigned_sections = cursor.fetchall()

    students = []
    selected_course = None
    selected_section_number = None  # Changed to be consistent
    selected_semester = None

    if request.method == 'POST':
        # ... [your existing POST handling code remains the same] ...
        pass
    elif request.args.get('course_id') and request.args.get('section_number'):
        selected_course = request.args.get('course_id')
        selected_section_number = request.args.get('section_number')
        
        # Find the matching section from assigned_sections
        section_info = None
        for section in assigned_sections:
            if (str(section['course_id']) == selected_course and 
                str(section['section_number']) == selected_section_number):
                section_info = section
                break
        
        if not section_info:
            return redirect(url_for('teacher_grades'))
            
        selected_semester = section_info['semester_id']
        section_id = section_info['section_id']

        # Get students enrolled in this section
        cursor.execute("""
            SELECT e.student_id, u.first_name, u.last_name
            FROM enrollments e
            JOIN students s ON s.student_id = e.student_id
            JOIN user u ON u.id = s.user_id
            WHERE e.section_id = %s
        """, (section_id,))
        students = cursor.fetchall()

    return render_template('teacher/grades.html', 
                         assigned_sections=assigned_sections, 
                         students=students,
                         selected_course=selected_course,
                         selected_section=selected_section_number)  # Pass section_number to template

@app.route('/teacher/announcements', methods=['GET', 'POST'])
def teacher_announcements():
    if 'loggedin' in session and session['role'] == 'teacher':
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # Get the teacher's sections
        cursor.execute("""
            SELECT s.section_number, c.title AS course_title
            FROM sections s
            JOIN courses c ON s.course_id = c.course_id
            WHERE s.instructor_id = %s
        """, (session['userid'],))
        sections = cursor.fetchall()

        # Handle announcement creation via POST
        if request.method == 'POST':
            title = request.form['title']
            message = request.form['message']
            section_id = request.form['section_id']

            cursor.execute("""
                INSERT INTO announcements (section_id, teacher_id, title, message)
                VALUES (%s, %s, %s, %s)
            """, (section_id, session['userid'], title, message))
            mysql.connection.commit()
            flash("Announcement created successfully!", 'success')
            return redirect(url_for('teacher_announcements'))

        # Get the teacher's announcements
        cursor.execute("""
            SELECT a.announcement_id, a.title, a.message, a.created_at, s.section_id, c.title AS course_title
            FROM announcements a
            JOIN sections s ON a.section_id = s.section_id
            JOIN courses c ON s.course_id = c.course_id
            WHERE a.teacher_id = %s
            ORDER BY a.created_at DESC
        """, (session['userid'],))
        announcements = cursor.fetchall()

        cursor.close()

        return render_template("teacher/announcements.html", sections=sections, announcements=announcements)

    return redirect(url_for('login'))


@app.route('/teacher/delete_announcement/<int:announcement_id>', methods=['POST'])
def delete_announcement(announcement_id):
    if 'loggedin' in session and session['role'] == 'teacher':
        cursor = mysql.connection.cursor()
        
        # Delete the announcement
        cursor.execute("DELETE FROM announcements WHERE announcement_id = %s AND teacher_id = %s", 
                       (announcement_id, session['userid']))
        mysql.connection.commit()
        cursor.close()
        
        flash("Announcement deleted successfully!", 'success')
        return redirect(url_for('teacher_announcements'))
    
    return redirect(url_for('login'))


@app.route('/teacher/exam_results', methods=['GET', 'POST'])
def teacher_exam_results():
    if 'loggedin' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    teacher_user_id = session['userid']
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Get teacher_id
    cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (teacher_user_id,))
    teacher = cursor.fetchone()
    if not teacher:
        return "Teacher profile not found."
    teacher_id = teacher['teacher_id']

    # Get assigned courses and sections
    cursor.execute("""
        SELECT c.course_id, c.title, s.section_number
        FROM course_section_teacher cst
        JOIN courses c ON c.course_id = cst.course_id
        JOIN sections s ON s.section_id = cst.section_id
        WHERE cst.teacher_id = %s
    """, (teacher_id,))
    assignments = cursor.fetchall()

    students_results = []
    
    if request.method == 'POST':
        course_id = request.form.get('course')
        section_id = request.form.get('section')

        if course_id and section_id:
            # Get enrolled students and their final marks
            cursor.execute("""
                SELECT e.student_id, u.first_name, u.last_name, er.final_result
                FROM enrollments e
                JOIN students s ON e.student_id = s.student_id
                JOIN user u ON s.user_id = u.id
                LEFT JOIN exam_results er ON er.student_id = e.student_id 
                    AND er.course_id = e.course_id
                WHERE e.section_id = %s AND e.course_id = %s
            """, (section_id, course_id))
            students_results = cursor.fetchall()

    return render_template(
        'teacher/exam_results.html',
        assignments=assignments,
        students_results=students_results,
        selected_course=request.form.get('course'),
        selected_section=request.form.get('section')
    )


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/teacher/assignments', methods=['GET', 'POST'])
def teacher_assignments():
    if 'loggedin' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Get teacher ID
    cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (session['userid'],))
    teacher = cursor.fetchone()
    if not teacher:
        flash("Teacher record not found.", "danger")
        return redirect(url_for('dashboard'))

    teacher_id = teacher['teacher_id']

    # Get current semester
    cursor.execute("SELECT * FROM semesters WHERE is_current = TRUE LIMIT 1")
    current_semester = cursor.fetchone()

    # Get assigned sections
    cursor.execute("""
        SELECT 
        c.title AS course_title,
        s.section_number,
        s.semester_id
        FROM course_section_teacher cst
        JOIN courses c ON c.course_id = cst.course_id
        JOIN sections s ON s.section_id = cst.section_id
        WHERE cst.teacher_id = %s AND cst.semester_id = %s
    """, (teacher_id, current_semester['semester_id'] if current_semester else None))
    sections = cursor.fetchall()

    # Handle submission
    if request.method == 'POST':
        section_id = request.form.get('section_id')
        title = request.form.get('title')
        description = request.form.get('description')
        due_date = request.form.get('due_date')
        max_points = request.form.get('max_points')

        # Handle file
        uploaded_file = request.files.get('file')
        file_path = None
        if uploaded_file and allowed_file(uploaded_file.filename):
            filename = secure_filename(uploaded_file.filename)
            save_path = os.path.join(app.config['UPLOAD_FOLDER2'], filename)
            os.makedirs(app.config['UPLOAD_FOLDER2'], exist_ok=True)
            uploaded_file.save(save_path)
            file_path = save_path
        else:
            file_path = None
        cursor.execute("""
            INSERT INTO assignments (section_id, title, description, due_date, max_points, file_path)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (section_id, title, description, due_date, max_points, file_path))
        mysql.connection.commit()
        flash('Assignment created successfully.', 'success')
        return redirect(url_for('teacher_assignments'))

    # Fetch assignments
    section_ids = [s['section_number'] for s in sections]
    assignments = []
    if section_ids:
        format_ids = ','.join(['%s'] * len(section_ids))
        cursor.execute(f"""
            SELECT a.*, c.title AS course_title
            FROM assignments a
            JOIN sections s ON a.section_id = s.section_number  
            JOIN courses c ON s.course_id = c.course_id
            WHERE a.section_id IN ({format_ids})
            ORDER BY a.due_date ASC
        """, tuple(section_ids))
        assignments = cursor.fetchall()

    cursor.close()

    return render_template("teacher/assignments.html", assignments=assignments, sections=sections)

@app.route('/teacher/assignments/edit/<int:assignment_id>', methods=['GET', 'POST'])
def edit_assignment(assignment_id):
    # Check if the user is logged in and is a teacher
    if 'loggedin' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    # Get teacher_id
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (session['userid'],))
    teacher = cursor.fetchone()
    if not teacher:
        return "Teacher profile not found.", 403
    teacher_id = teacher['teacher_id']

    # Get assignment details based on assignment_id
    cursor.execute("SELECT * FROM assignments WHERE assignment_id = %s", (assignment_id,))
    assignment = cursor.fetchone()

    if not assignment:
        return "Assignment not found", 404

    # If the form is submitted via POST
    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        due_date = request.form['due_date']
        max_points = request.form['max_points']
        uploaded_file = request.files.get('file')

        # Handle file upload if there is one
        if uploaded_file:
            filename = secure_filename(uploaded_file.filename)
            save_path = os.path.join(app.config['UPLOAD_FOLDER2'], filename)
            os.makedirs(app.config['UPLOAD_FOLDER2'], exist_ok=True)
            uploaded_file.save(save_path)
            file_path = save_path
        else:
            file_path = assignment['file_path']  # Keep existing file if no new one

        # Update the assignment in the database
        cursor.execute("""
            UPDATE assignments
            SET title = %s, description = %s, due_date = %s, max_points = %s, file_path = %s
            WHERE assignment_id = %s
        """, (title, description, due_date, max_points, file_path, assignment_id))
        mysql.connection.commit()

        return redirect(url_for('teacher_assignments'))

    return render_template('teacher/edit_assignment.html', assignment=assignment)

@app.route('/teacher/office_hours', methods=['GET', 'POST'])
def teacher_office_hours():
    if 'loggedin' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    teacher_user_id = session['userid']
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Get teacher_id
    cursor.execute("SELECT teacher_id FROM teacher WHERE user_id = %s", (teacher_user_id,))
    teacher = cursor.fetchone()
    if not teacher:
        return "Teacher profile not found."

    teacher_id = teacher['teacher_id']

    # Fetch distinct assigned courses
    cursor.execute("""
        SELECT DISTINCT c.course_id, c.title
        FROM course_section_teacher cst
        JOIN courses c ON c.course_id = cst.course_id
        WHERE cst.teacher_id = %s
    """, (teacher_id,))
    assigned_courses = cursor.fetchall()

    if request.method == 'POST':
        day = request.form['day']
        start_time = request.form['start_time']
        end_time = request.form['end_time']
        course_id = request.form['course_id']

        cursor.execute('''
            INSERT INTO office_hours (teacher_id, course_id, day, start_time, end_time)
            VALUES (%s, %s, %s, %s, %s)
        ''', (teacher_id, course_id, day, start_time, end_time))
        mysql.connection.commit()
        return redirect(url_for('teacher_office_hours'))

    # Fetch office hours with course titles
    cursor.execute('''
        SELECT o.day, o.start_time, o.end_time, c.title AS course_title
        FROM office_hours o
        JOIN courses c ON o.course_id = c.course_id
        WHERE o.teacher_id = %s
        ORDER BY FIELD(o.day, 'monday','tuesday','wednesday','thursday','friday'), o.start_time
    ''', (teacher_id,))
    office_hours = cursor.fetchall()

    return render_template('teacher/office_hours.html', office_hours=office_hours, courses=assigned_courses)

if __name__ == "__main__":
    app.run(debug=True)